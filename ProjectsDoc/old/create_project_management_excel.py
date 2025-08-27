import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, Reference
import os
from datetime import datetime, timedelta

def create_django_project_management_excel():
    """Django projesi için kapsamlı proje yönetimi Excel dosyası oluşturur"""
    
    wb = Workbook()
    
    # Varsayılan sheet'i sil
    default_sheet = wb.active
    if default_sheet:
        wb.remove(default_sheet)
    
    # Soft renk paletini tanımla (Pastel ve yumuşak tonlar)
    colors = {
        'primary': '7FB3D3',      # Soft Mavi
        'secondary': 'C39BD3',    # Soft Mor
        'success': '82E0AA',      # Soft Yeşil
        'warning': 'F7DC6F',      # Soft Sarı
        'danger': 'F1948A',       # Soft Kırmızı
        'info': '85C1E9',         # Soft Açık Mavi
        'light': 'F8F9FA',        # Çok Açık Gri
        'dark': '5D6D7E',         # Soft Koyu Gri
        'header1': '5DADE2',      # Soft Header Mavi
        'header2': 'BB8FCE',      # Soft Header Mor
        'header3': '58D68D',      # Soft Header Yeşil
        'header4': 'F8C471',      # Soft Header Turuncu
        'header5': 'EC7063',      # Soft Header Kırmızı
        'planning': 'AED6F1',     # Planlama - Açık Mavi
        'development': 'D5DBDB',  # Geliştirme - Gri
        'testing': 'FADBD8',      # Test - Açık Pembe
        'deployment': 'D5F4E6',   # Deployment - Açık Yeşil
        'maintenance': 'FCF3CF'   # Bakım - Açık Sarı
    }
    
    # Stil tanımları
    def create_header_style(color_name, text_color='FFFFFF'):
        style = NamedStyle(name=f"header_{color_name}")
        style.font = Font(bold=True, color=text_color, size=12)
        style.fill = PatternFill(start_color=colors[color_name], end_color=colors[color_name], fill_type="solid")
        style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        style.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        return style
    
    # Stilleri kaydet
    header_styles = {}
    for color in ['header1', 'header2', 'header3', 'header4', 'header5']:
        header_styles[color] = create_header_style(color)
        wb.add_named_style(header_styles[color])
    
    # 1. PROJE GENELİ SHEET
    overview_ws = wb.create_sheet("📋 Proje Genel")
    
    # Proje bilgileri
    overview_headers = [
        "📊 Proje Bilgisi", "💼 Değer", "📝 Notlar"
    ]
    
    for col, header in enumerate(overview_headers, 1):
        cell = overview_ws.cell(row=1, column=col, value=header)
        cell.style = header_styles['header1']
    
    # Proje bilgileri verileri
    project_info = [
        ["🎯 Proje Adı", "[PROJE ADI]", "Proje adını buraya yazın"],
        ["👤 Proje Yöneticisi", "[İSİM]", "Proje sorumlusunun adı"],
        ["📅 Başlama Tarihi", datetime.now().strftime('%d.%m.%Y'), "Projenin başlama tarihi"],
        ["📅 Hedef Bitiş", (datetime.now() + timedelta(days=90)).strftime('%d.%m.%Y'), "Hedeflenen bitiş tarihi"],
        ["💰 Bütçe", "[BÜTÇE]", "Proje bütçesi"],
        ["🎯 Hedef Kitle", "[HEDEf KİTLE]", "Uygulamanın hedef kitlesi"],
        ["🌐 Domain", "[DOMAIN.COM]", "Proje domain adresi"],
        ["🗄️ Veritabanı", "PostgreSQL", "Kullanılacak veritabanı"],
        ["☁️ Sunucu", "[SUNUCU BİLGİSİ]", "Hosting/sunucu bilgileri"],
        ["🔧 Django Versiyon", "5.0+", "Kullanılacak Django versiyonu"],
        ["🐍 Python Versiyon", "3.11+", "Kullanılacak Python versiyonu"],
        ["📱 Frontend", "[React/Vue/Angular]", "Frontend teknolojisi"],
        ["🚀 Deployment", "GitHub Actions", "Deployment yöntemi"],
        ["📊 Proje Durumu", "🟡 Planlama", "Mevcut proje durumu"],
        ["✅ Tamamlanma %", "0%", "Genel tamamlanma yüzdesi"]
    ]
    
    for row_num, info in enumerate(project_info, 2):
        for col_num, value in enumerate(info, 1):
            cell = overview_ws.cell(row=row_num, column=col_num, value=value)
            if col_num == 1:
                cell.font = Font(bold=True, color=colors['dark'])
                cell.fill = PatternFill(start_color=colors['light'], end_color=colors['light'], fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Sütun genişlikleri
    overview_ws.column_dimensions['A'].width = 25
    overview_ws.column_dimensions['B'].width = 20
    overview_ws.column_dimensions['C'].width = 35
    
    # 2. DETAYLI PROJE AŞAMALARI SHEET - Çok Kapsamlı
    phases_ws = wb.create_sheet("🎯 Detaylı Proje Aşamaları")
    
    phases_headers = [
        "🔢 Sıra", "📋 Aşama Adı", "📝 Detaylı Açıklama", "⏱️ Tahmini Süre", 
        "📊 Öncelik", "📅 Başlama Tarihi", "📅 Bitiş Tarihi", "✅ Durum", 
        "💡 İpuçları", "🔗 Bağımlılıklar", "📋 Notlar", "🎯 Kategori"
    ]
    
    for col, header in enumerate(phases_headers, 1):
        cell = phases_ws.cell(row=1, column=col, value=header)
        cell.style = header_styles['header2']
    
    # Çok detaylı proje aşamaları verileri
    detailed_phases = [
        # FAZE 1: PROJE PLANLAMA VE HAZIRLIK
        ["001", "📊 Proje Fikrini Netleştir", "Projenin tam olarak ne yapacağını, hangi problemi çözeceğini belirle. Hedef kitleyi tanımla.", "1-2 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Kağıt kalem ile brainstorming yap", "Yok", "Fikir netleşmeden başlama", "🎯 Planlama"],
        ["002", "🎯 Gereksinim Analizi Yap", "Projenin tüm fonksiyonel ve teknik gereksinimlerini listele. User story'leri yaz.", "2-3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Trello/Notion kullanarak organize ol", "001", "Eksik gereksinim sonradan büyük sorun", "🎯 Planlama"],
        ["003", "🗂️ Proje Klasör Yapısını Planla", "Django projesi için klasör hiyerarşisini kağıt üzerinde çiz", "0.5 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Django best practices araştır", "002", "", "🎯 Planlama"],
        ["004", "📊 Teknoloji Stack'i Belirle", "Django version, Python version, veritabanı, frontend teknolojilerini seç", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "LTS versiyonları tercih et", "002", "Güncel ve stabil sürümler seç", "🎯 Planlama"],
        ["005", "🗄️ Veritabanı Şeması Tasarla", "Tüm tabloları, ilişkileri, alanları kağıt/dijital olarak tasarla", "2-3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "draw.io veya Lucidchart kullan", "002,004", "Normalizasyon kurallarına dikkat", "🎯 Planlama"],
        ["006", "🎨 UI/UX Wireframe Hazırla", "Tüm sayfaların kabataslak çizimini yap (kağıt veya Figma)", "2-4 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Basit çizimlerle başla", "002", "Kullanıcı deneyimini öncelikle düşün", "🎯 Planlama"],
        ["007", "📋 Proje Zaman Çizelgesi Oluştur", "Her aşama için gerçekçi süre planlaması yap", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Buffer süre ekle", "001-006", "Çok iyimser tahmin yapma", "🎯 Planlama"],
        
        # FAZE 2: GELIŞTIRME ORTAMI HAZIRLIĞI
        ["008", "🐍 Python Kurulumu Kontrol Et", "Python 3.8+ kurulu olduğunu doğrula, gerekirse güncelle", "0.5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "python --version komutu çalıştır", "Yok", "Virtual environment öncesi şart", "⚙️ Kurulum"],
        ["009", "📦 IDE/Editor Seç ve Konfigüre Et", "VSCode, PyCharm vb. seç, Django extension'ları yükle", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Linting, formatting araçlarını kur", "008", "Kod yazma hızını artırır", "⚙️ Kurulum"],
        ["010", "📁 Ana Proje Klasörü Oluştur", "Bilgisayarda projenin ana klasörünü oluştur", "0.1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Türkçe karakter kullanma", "003", "", "⚙️ Kurulum"],
        ["011", "🔧 Virtual Environment Oluştur", "python -m venv venv komutu ile sanal ortam oluştur", "0.5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Her projede ayrı venv kullan", "008,010", "Global Python'u kirletme", "⚙️ Kurulum"],
        ["012", "🔄 Virtual Environment Aktif Et", "venv\\Scripts\\activate (Windows) komutu çalıştır", "0.1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Terminal'de (venv) görünmeli", "011", "Her çalışma öncesi aktif et", "⚙️ Kurulum"],
        ["013", "📦 Django Kurulumu", "pip install django komutu ile Django'yu yükle", "0.5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "pip install django==5.0 belirli versiyon", "012", "LTS versiyonu tercih et", "⚙️ Kurulum"],
        ["014", "🏗️ Django Projesi Oluştur", "django-admin startproject myproject komutu çalıştır", "0.5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Proje adında Türkçe karakter yok", "013", "Anlamlı isim seç", "⚙️ Kurulum"],
        ["015", "🎯 Django Apps Planla", "Projenin modülerlik için hangi app'lere bölüneceğini belirle", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Her özellik için ayrı app", "005,014", "Sonradan değiştirmek zor", "⚙️ Kurulum"],
        ["016", "📱 Django Apps Oluştur", "python manage.py startapp appname komutları çalıştır", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Tekil isim kullan (user, not users)", "015", "INSTALLED_APPS'e eklemeyi unutma", "⚙️ Kurulum"],
        ["017", "🗄️ PostgreSQL Kurulumu", "PostgreSQL'i indir, kur ve konfigüre et", "2 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "pgAdmin da kur", "004", "Varsayılan şifre not et", "⚙️ Kurulum"],
        ["018", "🔗 Veritabanı Bağlantısı Kur", "Django settings.py'da PostgreSQL konfigürasyonu yap", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "psycopg2-binary paketi gerekli", "017", "Bağlantı bilgilerini test et", "⚙️ Kurulum"],
        ["019", "📋 Requirements.txt Oluştur", "pip freeze > requirements.txt komutu çalıştır", "0.5 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Düzenli güncelle", "013,018", "Version control için önemli", "⚙️ Kurulum"],
        ["020", "🌍 Git Repository Oluştur", "git init, git add, git commit komutları çalıştır", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", ".gitignore dosyası ekle", "019", "Her önemli değişiklikte commit", "⚙️ Kurulum"],
        ["021", "🐙 GitHub'a Push Et", "GitHub'da repo oluştur ve push et", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "README.md dosyası ekle", "020", "Günlük push yapma alışkanlığı", "⚙️ Kurulum"],
        ["022", "🔧 Django Settings Yapılandır", "Development/Production ayarlarını düzenle", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Çevresel değişkenler kullan", "018", "SECRET_KEY'i güvenli tut", "⚙️ Kurulum"],
        
        # FAZE 3: MODEL TASARIMI VE VERİTABANI
        ["023", "📊 Model Sınıfları Oluştur", "Her app için models.py dosyalarında model sınıflarını yaz", "3-5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Field tiplerini doğru seç", "005,016", "Meta sınıflarını unutma", "🗄️ Veritabanı"],
        ["024", "🔗 Model İlişkileri Kur", "ForeignKey, ManyToMany, OneToOne ilişkilerini tanımla", "2-3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "related_name kullan", "023", "Circular import'a dikkat", "🗄️ Veritabanı"],
        ["025", "✅ Model Validation Ekle", "clean() metodları ve field validationları ekle", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Django validators kullan", "024", "Veri tutarlılığı için kritik", "🗄️ Veritabanı"],
        ["026", "📝 Model __str__ Metodları", "Her model için anlamlı __str__ metodu yaz", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Admin panelde görünür", "023", "Debug'da çok yardımcı", "🗄️ Veritabanı"],
        ["027", "🗂️ İlk Migration Oluştur", "python manage.py makemigrations komutu çalıştır", "0.5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Migration dosyalarını kontrol et", "025", "Git'e commit et", "🗄️ Veritabanı"],
        ["028", "⚡ Migration Uygula", "python manage.py migrate komutu çalıştır", "0.5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Hata mesajlarını dikkatle oku", "027", "Database backup al", "🗄️ Veritabanı"],
        ["029", "👑 Superuser Oluştur", "python manage.py createsuperuser komutu çalıştır", "0.5 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Güçlü şifre kullan", "028", "Admin paneline giriş için", "🗄️ Veritabanı"],
        ["030", "🔧 Admin Panel Konfigürasyonu", "admin.py dosyalarında model kayıtlarını yap", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "list_display, search_fields ekle", "029", "Veri girişi kolaylaşır", "🗄️ Veritabanı"],
        ["031", "📊 Sample Data Ekleme", "Fixture'lar veya management command ile test verisi ekle", "1-2 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Gerçekçi veriler kullan", "030", "Development'ta çok yardımcı", "🗄️ Veritabanı"],
        
        # FAZE 4: VIEWS VE URL YAPILANDIRMA
        ["032", "🗺️ URL Patterns Planla", "Tüm sayfa URL'lerini ve API endpoint'lerini planla", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "RESTful naming convention", "006", "SEO dostu URL'ler", "🌐 Backend"],
        ["033", "🏠 Ana Sayfa View Oluştur", "Home page için view fonksiyonu/sınıfı yaz", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "TemplateView veya function view", "032", "İlk çalışan sayfa", "🌐 Backend"],
        ["034", "🔗 URL Configuration", "URLconf'ları (urls.py) yapılandır", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "App seviyesinde URL'ler", "033", "name parametreleri ekle", "🌐 Backend"],
        ["035", "📋 CRUD Views Oluştur", "Her model için Create, Read, Update, Delete view'ları", "5-7 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Class-based views kullan", "031,034", "Generic views ile hızlandır", "🌐 Backend"],
        ["036", "🔐 Authentication Views", "Login, logout, register view'larını oluştur", "2-3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Django'nun hazır view'larını kullan", "035", "Password reset de ekle", "🌐 Backend"],
        ["037", "🛡️ Permission Kontrolü", "View'larda izin kontrollerini ekle", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "login_required decorator", "036", "Security için kritik", "🌐 Backend"],
        ["038", "📊 Pagination Ekle", "Uzun listelerde sayfa numaralandırma", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Django Paginator kullan", "035", "Performance için önemli", "🌐 Backend"],
        ["039", "🔍 Search Functionality", "Arama özelliği ekle", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Q objects kullan", "035", "User experience", "🌐 Backend"],
        ["040", "📱 API Views (Opsiyonel)", "REST API için DRF view'ları", "3-5 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Django REST Framework", "035", "Mobile app için gerekebilir", "🌐 Backend"],
        
        # FAZE 5: TEMPLATES VE FRONTEND
        ["041", "📁 Template Klasör Yapısı", "templates klasörünü ve alt klasörlerini oluştur", "0.5 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "App seviyesinde organize et", "006", "Karışıklığı önler", "🎨 Frontend"],
        ["042", "🏗️ Base Template Oluştur", "base.html ana şablon dosyasını oluştur", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Block'ları doğru tanımla", "041", "Tüm sayfalar bunu extend eder", "🎨 Frontend"],
        ["043", "🎨 CSS Framework Entegrasyonu", "Bootstrap, Tailwind vb. framework ekle", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "CDN veya lokal kurulum", "042", "Responsive tasarım için", "🎨 Frontend"],
        ["044", "🏠 Ana Sayfa Template", "index.html ana sayfa şablonunu oluştur", "2 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "SEO meta tag'ları ekle", "033,043", "İlk izlenim çok önemli", "🎨 Frontend"],
        ["045", "📝 Form Templates", "Tüm form sayfalarının template'lerini oluştur", "3-4 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "CSRF token'ı unutma", "035,044", "User input için gerekli", "🎨 Frontend"],
        ["046", "📋 List Templates", "Listeleme sayfalarının template'lerini oluştur", "2-3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Pagination template ekle", "035,038", "Veri görüntüleme", "🎨 Frontend"],
        ["047", "📄 Detail Templates", "Detay sayfalarının template'lerini oluştur", "2 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Breadcrumb ekle", "035,046", "İçerik detayları", "🎨 Frontend"],
        ["048", "🔐 Auth Templates", "Login, register template'lerini oluştur", "2 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Error message'ları ekle", "036,045", "Kullanıcı işlemleri", "🎨 Frontend"],
        ["049", "📱 Responsive Design", "Mobil uyumlu tasarım yap", "3-4 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Media queries kullan", "043-048", "Mobil kullanım artıyor", "🎨 Frontend"],
        ["050", "⚡ JavaScript Functionality", "İnteraktif özellikler ekle", "2-3 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "jQuery veya Vanilla JS", "049", "User experience artırır", "🎨 Frontend"],
        
        # FAZE 6: FORMS VE VALİDASYON
        ["051", "📝 Django Forms Oluştur", "Her model için form sınıfları oluştur", "2-3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "ModelForm kullan", "025,045", "Validation için gerekli", "📝 Forms"],
        ["052", "✅ Form Validation", "Form seviyesinde validation kuralları ekle", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "clean() metodları yaz", "051", "Veri kalitesi için", "📝 Forms"],
        ["053", "🎨 Form Styling", "Form'ları güzel görünüm için stil ver", "1-2 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Widget'ları özelleştir", "052", "User experience", "📝 Forms"],
        ["054", "💾 File Upload Forms", "Dosya yükleme form'larını ekle", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "FileField/ImageField", "051", "Media dosyalar için", "📝 Forms"],
        ["055", "🔒 CSRF Protection", "Tüm form'larda CSRF koruması kontrol et", "0.5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "{% csrf_token %} template'lerde", "045", "Security açığı önler", "📝 Forms"],
        
        # FAZE 7: STATIC FILES VE MEDIA
        ["056", "📁 Static Files Yapılandırma", "CSS, JS, images için static klasör yapısı", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "App seviyesinde static klasör", "022", "Development/Production farkı", "📁 Static"],
        ["057", "🎨 CSS Dosyaları Oluştur", "Custom CSS stillerini yaz", "2-3 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Modüler CSS yaz", "056", "Brand identity için", "📁 Static"],
        ["058", "⚡ JavaScript Dosyaları", "Custom JavaScript kodlarını yaz", "2 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "ES6 syntax kullan", "057", "İnteraktivite için", "📁 Static"],
        ["059", "🖼️ Media Files Konfigürasyonu", "Kullanıcı yükleme dosyaları için media ayarları", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "MEDIA_URL ve MEDIA_ROOT", "054", "File upload için gerekli", "📁 Static"],
        ["060", "🗜️ Static Files Optimize", "CSS/JS dosyalarını minimize et", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Django Compressor", "058", "Performance için", "📁 Static"],
        
        # FAZE 8: GÜVENLİK VE PERMİSSİONS
        ["061", "🔐 User Permission System", "Kullanıcı yetki sistemi kur", "3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Django Groups kullan", "037", "Role-based access", "🔒 Security"],
        ["062", "🛡️ Security Settings", "Production güvenlik ayarlarını yap", "2 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "HTTPS, SECURE_* settings", "022", "Canlıda mutlaka gerekli", "🔒 Security"],
        ["063", "⚡ Rate Limiting", "Brute force saldırılarına karşı koruma", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "django-ratelimit", "036", "API koruma", "🔒 Security"],
        ["064", "🔍 SQL Injection Koruması", "ORM kullanımını kontrol et", "0.5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Raw SQL kullanma", "035", "Güvenlik açığı", "🔒 Security"],
        ["065", "🍪 Session Security", "Oturum güvenliği ayarları", "0.5 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "SESSION_COOKIE_* settings", "062", "Oturum hijacking önleme", "🔒 Security"],
        
        # FAZE 9: TEST YAZMA
        ["066", "🧪 Unit Test Kurulumu", "Test ortamını ve framework'ünü hazırla", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Django TestCase kullan", "031", "Test coverage için", "🧪 Testing"],
        ["067", "📊 Model Tests", "Model sınıfları için test yaz", "2-3 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Edge case'leri test et", "066", "Veri bütünlüğü", "🧪 Testing"],
        ["068", "🌐 View Tests", "View fonksiyonları için test yaz", "3-4 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Response status code'ları", "067", "HTTP davranışları", "🧪 Testing"],
        ["069", "📝 Form Tests", "Form validation'ları test et", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Valid/invalid input test", "068", "Input validation", "🧪 Testing"],
        ["070", "🔐 Authentication Tests", "Login/logout süreçlerini test et", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Permission test'leri", "069", "Güvenlik testi", "🧪 Testing"],
        ["071", "📊 Test Coverage Analizi", "Test coverage raporunu çıkar", "0.5 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "coverage.py kullan", "070", "%80+ coverage hedefle", "🧪 Testing"],
        
        # FAZE 10: PERFORMANce OPTİMİZASYONU
        ["072", "🗄️ Database Optimization", "Veritabanı sorgularını optimize et", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Django Debug Toolbar", "071", "N+1 problem çöz", "⚡ Performance"],
        ["073", "📊 Index Ekleme", "Veritabanı indexlerini ekle", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "db_index=True", "072", "Query performance", "⚡ Performance"],
        ["074", "💾 Caching Sistemi", "Cache framework'ünü entegre et", "2 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Redis veya Memcached", "073", "Response time iyileştir", "⚡ Performance"],
        ["075", "🗜️ File Compression", "CSS/JS dosyalarını sıkıştır", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Whitenoise kullan", "060", "Load time azalt", "⚡ Performance"],
        ["076", "🖼️ Image Optimization", "Resim dosyalarını optimize et", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Pillow ile resize", "075", "Bandwidth tasarrufu", "⚡ Performance"],
        
        # FAZE 11: DEPLOYMENT HAZIRLIĞI
        ["077", "🐳 Docker Configuration", "Docker container'ı hazırla", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Multi-stage build", "076", "Consistent deployment", "🚀 Deployment"],
        ["078", "☁️ Production Settings", "Canlı ortam ayarlarını hazırla", "2 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Environment variables", "062", "Security için ayrı ayar", "🚀 Deployment"],
        ["079", "📊 Database Migration", "Production migration planını yap", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Data migration da düşün", "078", "Downtime minimize et", "🚀 Deployment"],
        ["080", "🔧 CI/CD Pipeline", "GitHub Actions workflow'unu yaz", "3 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Test → Build → Deploy", "021,071", "Otomatik deployment", "🚀 Deployment"],
        ["081", "🌐 Domain & SSL", "Domain satın al ve SSL sertifikası kur", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Let's Encrypt ücretsiz", "080", "HTTPS zorunlu", "🚀 Deployment"],
        ["082", "☁️ Server Setup", "Production sunucusunu hazırla", "2-3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "DigitalOcean, AWS vb.", "081", "Resource planning yap", "🚀 Deployment"],
        ["083", "🚀 First Deployment", "İlk canlı deployment'ı yap", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Staging ortamında test et", "082", "Rollback planı hazır olsun", "🚀 Deployment"],
        
        # FAZE 12: MONİTORING VE LOGGING
        ["084", "📊 Logging Setup", "Uygulama log'larını yapılandır", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Python logging module", "083", "Debug için gerekli", "📊 Monitoring"],
        ["085", "⚡ Performance Monitoring", "Uygulama performansını izle", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Sentry, New Relic", "084", "Real-time monitoring", "📊 Monitoring"],
        ["086", "🔔 Error Tracking", "Hata takip sistemi kur", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Sentry entegrasyonu", "085", "Proaktif hata çözme", "📊 Monitoring"],
        ["087", "💾 Backup System", "Otomatik backup sistemi kur", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Database + media files", "086", "Data loss önleme", "📊 Monitoring"],
        ["088", "📈 Analytics", "Kullanıcı davranışları analizi", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Google Analytics", "087", "User insights", "📊 Monitoring"],
        
        # FAZE 13: DOKÜMANTASYON
        ["089", "📚 API Documentation", "API endpoint'lerini dokümante et", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Swagger/OpenAPI", "040", "Developer experience", "📚 Documentation"],
        ["090", "📖 User Manual", "Son kullanıcı kılavuzu yaz", "2-3 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Screenshot'lar ekle", "089", "User adoption", "📚 Documentation"],
        ["091", "🔧 Technical Documentation", "Teknik dokümantasyon yaz", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Code comment'ları", "090", "Maintenance için", "📚 Documentation"],
        ["092", "🚀 Deployment Guide", "Deployment talimatları yaz", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Step-by-step guide", "091", "DevOps için", "📚 Documentation"],
        ["093", "📋 README.md Güncelle", "GitHub README'yi detaylandır", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Installation guide", "092", "Open source için önemli", "📚 Documentation"],
        
        # FAZE 14: TEST VE QA
        ["094", "🧪 Manual Testing", "Tüm özellikleri manuel olarak test et", "3-4 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Test case'leri hazırla", "071", "User acceptance test", "✅ QA"],
        ["095", "📱 Cross-browser Testing", "Farklı browser'larda test et", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Chrome, Firefox, Safari", "094", "Compatibility sağla", "✅ QA"],
        ["096", "📱 Mobile Testing", "Mobil cihazlarda test et", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Responsive tasarım kontrol", "095", "Mobile usage artıyor", "✅ QA"],
        ["097", "⚡ Load Testing", "Yüksek trafik altında test et", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Apache Bench kullan", "096", "Performance limit'leri", "✅ QA"],
        ["098", "🔒 Security Testing", "Güvenlik açıklarını test et", "2 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "OWASP Top 10", "097", "Security vulnerabilities", "✅ QA"],
        ["099", "🐛 Bug Fixes", "Testte bulunan hataları düzelt", "3-5 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Priority sıralaması yap", "098", "Quality assurance", "✅ QA"],
        
        # FAZE 15: CANLI YAYIN VE TAMAMLAMA
        ["100", "🚀 Final Deployment", "Son hali ile canlı yayın", "1 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Backup aldıktan sonra", "099", "Go-live moment", "🎯 Launch"],
        ["101", "📊 Post-launch Monitoring", "Canlı yayın sonrası takip", "3 gün", "🔴 Kritik", "", "", "⚪ Bekliyor", "Error rate'leri izle", "100", "Immediate issue response", "🎯 Launch"],
        ["102", "📈 Performance Review", "Performance metrikleri analiz et", "2 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Response time, load", "101", "Optimization insights", "🎯 Launch"],
        ["103", "📝 Project Retrospective", "Proje retrospektifi yap", "1 gün", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "Lessons learned", "102", "Future project improvement", "🎯 Launch"],
        ["104", "🎉 Project Completion", "Proje tamamlama dokümantasyonu", "1 gün", "🟡 Önemli", "", "", "⚪ Bekliyor", "Final report hazırla", "103", "Official closure", "🎯 Launch"],
        
        # FAZE 16: BAKIM VE GÜNCELLEMe
        ["105", "🔄 Regular Updates", "Düzenli güncelleme planı", "Sürekli", "🟡 Önemli", "", "", "⚪ Bekliyor", "Security patch'leri", "104", "System maintenance", "🔧 Maintenance"],
        ["106", "📊 Analytics Review", "Kullanım analitikleri inceleme", "Aylık", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "User behavior patterns", "105", "Feature planning", "🔧 Maintenance"],
        ["107", "🆕 Feature Requests", "Yeni özellik talepleri değerlendirme", "Sürekli", "🟢 İyi Olur", "", "", "⚪ Bekliyor", "User feedback priority", "106", "Product evolution", "🔧 Maintenance"],
        ["108", "🔒 Security Audits", "Düzenli güvenlik denetimleri", "3 Aylık", "🟡 Önemli", "", "", "⚪ Bekliyor", "Vulnerability scanning", "107", "Security maintenance", "🔧 Maintenance"],
    ]
    
    for row_num, phase in enumerate(detailed_phases, 2):
        for col_num, value in enumerate(phase, 1):
            cell = phases_ws.cell(row=row_num, column=col_num, value=value)
            
            # Kategori bazlı renklendirme
            category = phase[11] if len(phase) > 11 else ""
            if col_num == 12:  # Kategori sütunu
                if "Planlama" in category:
                    cell.fill = PatternFill(start_color=colors['planning'], end_color=colors['planning'], fill_type="solid")
                elif "Kurulum" in category:
                    cell.fill = PatternFill(start_color=colors['development'], end_color=colors['development'], fill_type="solid")
                elif "Veritabanı" in category:
                    cell.fill = PatternFill(start_color=colors['header1'], end_color=colors['header1'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Backend" in category:
                    cell.fill = PatternFill(start_color=colors['header2'], end_color=colors['header2'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Frontend" in category:
                    cell.fill = PatternFill(start_color=colors['header3'], end_color=colors['header3'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Testing" in category:
                    cell.fill = PatternFill(start_color=colors['testing'], end_color=colors['testing'], fill_type="solid")
                elif "Deployment" in category:
                    cell.fill = PatternFill(start_color=colors['deployment'], end_color=colors['deployment'], fill_type="solid")
                elif "Maintenance" in category:
                    cell.fill = PatternFill(start_color=colors['maintenance'], end_color=colors['maintenance'], fill_type="solid")
            
            # Öncelik renklendirme
            elif col_num == 5:  # Öncelik sütunu
                if "Kritik" in str(value):
                    cell.fill = PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Önemli" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='000000', bold=True)
                elif "İyi Olur" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
            
            # Durum renklendirme
            elif col_num == 8:  # Durum sütunu
                if "Tamamlandı" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Devam" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='000000', bold=True)
                elif "Bekliyor" in str(value):
                    cell.fill = PatternFill(start_color=colors['light'], end_color=colors['light'], fill_type="solid")
            
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Sütun genişlikleri
    column_widths_detailed = [8, 25, 40, 12, 12, 12, 12, 15, 35, 20, 30, 15]
    for i, width in enumerate(column_widths_detailed, 1):
        phases_ws.column_dimensions[get_column_letter(i)].width = width
    
    # 3. DETAYLI GÖREVLER SHEET
    tasks_ws = wb.create_sheet("📋 Detaylı Görevler")
    
    tasks_headers = [
        "🎯 Aşama", "📋 Görev ID", "📝 Görev Adı", "📄 Açıklama", "👤 Atanan", 
        "📅 Başlama", "📅 Bitiş", "⏱️ Süre", "📊 Durum", "✅ %", "🎯 Öncelik", "🏷️ Etiketler", "📋 Notlar"
    ]
    
    for col, header in enumerate(tasks_headers, 1):
        cell = tasks_ws.cell(row=1, column=col, value=header)
        cell.style = header_styles['header3']
    
    # Detaylı görevler verileri
    tasks_data = [
        # Proje Planlama görevleri
        ["1️⃣ Planlama", "T001", "📊 Gereksinim Analizi", "Müşteri ihtiyaçlarını belirle", "Analist", "01.09.2024", "03.09.2024", "3 gün", "🟢 Tamamlandı", "100%", "🔴 Yüksek", "analiz,gereksinim", "Müşteri onayı alındı"],
        ["1️⃣ Planlama", "T002", "📋 Proje Planı", "Detaylı proje planı hazırla", "PM", "04.09.2024", "06.09.2024", "3 gün", "🟡 Devam", "70%", "🔴 Yüksek", "planlama", "Gantt şeması tamamlandı"],
        ["1️⃣ Planlama", "T003", "👥 Ekip Oluşturma", "Proje ekibini belirle", "PM", "07.09.2024", "07.09.2024", "1 gün", "⚪ Bekliyor", "0%", "🟡 Orta", "ekip,kaynak", ""],
        
        # Sistem Tasarımı görevleri
        ["2️⃣ Tasarım", "T004", "🏗️ Sistem Mimarisi", "Genel sistem mimarisini tasarla", "Mimar", "08.09.2024", "10.09.2024", "3 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "mimari,tasarım", ""],
        ["2️⃣ Tasarım", "T005", "🎨 UI/UX Tasarım", "Kullanıcı arayüzü tasarımı", "Designer", "11.09.2024", "14.09.2024", "4 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "ui,ux,tasarım", "Figma kullanılacak"],
        
        # Veritabanı Tasarımı
        ["3️⃣ Veritabanı", "T006", "🗄️ Model Tasarımı", "Django modellerini tasarla", "Backend Dev", "15.09.2024", "16.09.2024", "2 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "model,django", "ERD diyagramı hazırla"],
        ["3️⃣ Veritabanı", "T007", "🔗 İlişki Tasarımı", "Model ilişkilerini belirle", "Backend Dev", "17.09.2024", "18.09.2024", "2 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "ilişki,foreign", ""],
        
        # Backend Geliştirme
        ["4️⃣ Backend", "T008", "🏗️ Proje Kurulumu", "Django projesi kur", "Backend Dev", "19.09.2024", "19.09.2024", "1 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "django,setup", "Virtual env kur"],
        ["4️⃣ Backend", "T009", "📝 Model Oluşturma", "Django modellerini oluştur", "Backend Dev", "20.09.2024", "22.09.2024", "3 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "model,oluştur", "Migration dosyaları"],
        ["4️⃣ Backend", "T010", "🔐 Authentication", "Kullanıcı kimlik doğrulama", "Backend Dev", "23.09.2024", "25.09.2024", "3 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "auth,jwt", "JWT token kullan"],
        ["4️⃣ Backend", "T011", "🌐 API Endpoints", "REST API endpoint'leri", "Backend Dev", "26.09.2024", "05.10.2024", "10 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "api,rest", "DRF kullan"],
        ["4️⃣ Backend", "T012", "📋 Admin Panel", "Django admin özelleştir", "Backend Dev", "06.10.2024", "08.10.2024", "3 gün", "⚪ Bekliyor", "0%", "🟡 Orta", "admin,panel", ""],
        ["4️⃣ Backend", "T013", "🔍 API Dokümantasyon", "Swagger/OpenAPI docs", "Backend Dev", "09.10.2024", "10.10.2024", "2 gün", "⚪ Bekliyor", "0%", "🟡 Orta", "docs,swagger", ""],
        
        # Frontend görevleri
        ["5️⃣ Frontend", "T014", "⚛️ React Kurulumu", "React projesi kur", "Frontend Dev", "01.10.2024", "02.10.2024", "2 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "react,setup", "TypeScript kullan"],
        ["5️⃣ Frontend", "T015", "🎨 Component Yapısı", "Temel component'leri oluştur", "Frontend Dev", "03.10.2024", "08.10.2024", "6 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "component,react", ""],
        ["5️⃣ Frontend", "T016", "📱 Responsive Tasarım", "Mobil uyumlu tasarım", "Frontend Dev", "09.10.2024", "15.10.2024", "7 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "responsive,mobile", "Bootstrap/Tailwind"],
        ["5️⃣ Frontend", "T017", "🔗 API Entegrasyonu", "Backend API'leri bağla", "Frontend Dev", "16.10.2024", "20.10.2024", "5 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "api,axios", ""],
        ["5️⃣ Frontend", "T018", "🎯 State Management", "Redux/Context kurulumu", "Frontend Dev", "21.10.2024", "25.10.2024", "5 gün", "⚪ Bekliyor", "0%", "🟡 Orta", "redux,state", ""],
        
        # Test görevleri
        ["7️⃣ Test", "T019", "🧪 Unit Testler", "Backend unit testleri", "QA", "01.11.2024", "05.11.2024", "5 gün", "⚪ Bekliyor", "0%", "🟡 Orta", "test,unit", "pytest kullan"],
        ["7️⃣ Test", "T020", "🔗 Integration Test", "API integration testleri", "QA", "06.11.2024", "08.11.2024", "3 gün", "⚪ Bekliyor", "0%", "🟡 Orta", "test,integration", ""],
        ["7️⃣ Test", "T021", "👤 User Acceptance", "Kullanıcı kabul testleri", "QA", "09.11.2024", "10.11.2024", "2 gün", "⚪ Bekliyor", "0%", "🟢 Düşük", "test,uat", "Müşteri ile birlikte"],
        
        # Deployment
        ["8️⃣ Deployment", "T022", "🐳 Docker Setup", "Docker container oluştur", "DevOps", "11.11.2024", "12.11.2024", "2 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "docker,container", ""],
        ["8️⃣ Deployment", "T023", "🚀 CI/CD Pipeline", "GitHub Actions kur", "DevOps", "13.11.2024", "14.11.2024", "2 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "cicd,github", "Otomatik deployment"],
        ["8️⃣ Deployment", "T024", "☁️ Cloud Deploy", "Canlı sunucuya deploy", "DevOps", "15.11.2024", "15.11.2024", "1 gün", "⚪ Bekliyor", "0%", "🔴 Yüksek", "cloud,deploy", ""],
    ]
    
    for row_num, task in enumerate(tasks_data, 2):
        for col_num, value in enumerate(task, 1):
            cell = tasks_ws.cell(row=row_num, column=col_num, value=value)
            
            # Durum renklendirme
            if col_num == 9:  # Durum sütunu
                if "Tamamlandı" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Devam" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Bekliyor" in str(value):
                    cell.fill = PatternFill(start_color=colors['light'], end_color=colors['light'], fill_type="solid")
            
            elif col_num == 11:  # Öncelik
                if "Yüksek" in str(value):
                    cell.fill = PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Orta" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Düşük" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
            
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Sütun genişlikleri
    column_widths_tasks = [15, 10, 25, 30, 15, 12, 12, 8, 15, 8, 12, 20, 25]
    for i, width in enumerate(column_widths_tasks, 1):
        tasks_ws.column_dimensions[get_column_letter(i)].width = width
    
    # 3. KİŞİSEL GELİŞTİRİCİ TAKİBİ SHEET
    developer_ws = wb.create_sheet("👨‍💻 Geliştirici Takibi")
    
    dev_headers = [
        "📅 Tarih", "⏰ Başlama Saati", "⏰ Bitiş Saati", "🎯 Yapılan İş", 
        "📝 Detay Açıklama", "⏱️ Süre (Saat)", "😊 Enerji Seviyesi", "🎯 Verimlilik", 
        "🐛 Karşılaşılan Problemler", "💡 Öğrenilenler", "📋 Yarın Planı", "📊 Not"
    ]
    
    for col, header in enumerate(dev_headers, 1):
        cell = developer_ws.cell(row=1, column=col, value=header)
        cell.style = header_styles['header4']
    
    # Geliştirici günlük takip örnekleri
    dev_data = [
        ["27.08.2024", "09:00", "17:00", "Proje planlaması ve gereksinim analizi", "User story'ler yazıldı, wireframe çizildi", "8", "😊 Yüksek", "⭐⭐⭐⭐⭐", "Veritabanı tasarımında kararsızlık", "PostgreSQL vs MySQL karşılaştırması", "Django projesini oluşturma", "Verimli geçti"],
        ["28.08.2024", "10:00", "18:00", "Django projesi kurulumu", "Virtual env, Django install, ilk app", "8", "😐 Orta", "⭐⭐⭐⭐", "Django ayarlarında konfüzyon", "Settings.py yapısını öğrendim", "Model tasarımına başlama", "Öğretici videolar izlendi"],
        ["29.08.2024", "", "", "", "", "", "", "", "", "", "", ""],
        ["30.08.2024", "", "", "", "", "", "", "", "", "", "", ""],
        ["31.08.2024", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    
    for row_num, dev_day in enumerate(dev_data, 2):
        for col_num, value in enumerate(dev_day, 1):
            cell = developer_ws.cell(row=row_num, column=col_num, value=value)
            
            # Enerji seviyesi renklendirme
            if col_num == 7:  # Enerji seviyesi
                if "Yüksek" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Orta" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='000000', bold=True)
                elif "Düşük" in str(value):
                    cell.fill = PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
            
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    
    # 4. GÜNLÜK GÖREV TAKİBİ SHEET
    daily_ws = wb.create_sheet("📅 Günlük Görev Takibi")
    
    daily_headers = [
        "📅 Tarih", "🎯 Ana Görev", "📋 Alt Görevler", "📊 Öncelik", "⏰ Tahmini Süre", 
        "✅ Tamamlandı mı?", "⏱️ Gerçek Süre", "🎯 Verimlilik %", "😊 Zorluk", "📝 Notlar", 
        "🐛 Problemler", "💡 Çözümler", "📚 Öğrenilenler"
    ]
    
    for col, header in enumerate(daily_headers, 1):
        cell = daily_ws.cell(row=1, column=col, value=header)
        cell.style = header_styles['header3']
    
    # Günlük görev örnekleri
    daily_data = [
        ["27.08.2024", "Django Kurulumu", "Virtual env oluştur, Django yükle, proje oluştur", "🔴 Yüksek", "4 saat", "✅ Evet", "3.5 saat", "90%", "😊 Kolay", "Hızlı tamamlandı", "Yok", "Terminal komutları", "Django project structure"],
        ["27.08.2024", "Model Tasarımı", "User, Product, Order modellerini planla", "🔴 Yüksek", "3 saat", "⚠️ Kısmi", "4 saat", "75%", "🤔 Orta", "İlişkiler karmaşık", "ForeignKey vs OneToOne", "Django docs", "Model relationships"],
        ["28.08.2024", "Models.py Oluşturma", "Tüm modelleri kod olarak yaz", "🔴 Yüksek", "2 saat", "✅ Evet", "2.5 saat", "80%", "😊 Kolay", "Syntax hatalar", "Field types", "Model validation", "Django ORM"],
        ["28.08.2024", "Migration İşlemleri", "makemigrations ve migrate", "🟡 Orta", "1 saat", "✅ Evet", "0.5 saat", "100%", "😊 Kolay", "Sorunsuz", "Yok", "Migration komutları", "Database schemas"],
        ["29.08.2024", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["29.08.2024", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["30.08.2024", "", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    
    for row_num, daily_task in enumerate(daily_data, 2):
        for col_num, value in enumerate(daily_task, 1):
            cell = daily_ws.cell(row=row_num, column=col_num, value=value)
            
            # Tamamlanma durumu renklendirme
            if col_num == 6:  # Tamamlandı mı
                if "Evet" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Kısmi" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='000000', bold=True)
                elif "Hayır" in str(value):
                    cell.fill = PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
            
            # Öncelik renklendirme
            elif col_num == 4:  # Öncelik
                if "Yüksek" in str(value):
                    cell.fill = PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Orta" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='000000', bold=True)
                elif "Düşük" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
            
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Sütun genişlikleri
    column_widths_daily = [12, 20, 35, 12, 12, 15, 12, 12, 12, 25, 25, 25, 25]
    for i, width in enumerate(column_widths_daily, 1):
        daily_ws.column_dimensions[get_column_letter(i)].width = width
    
    # 5. RİSK YÖNETİMİ SHEET
    risk_ws = wb.create_sheet("⚠️ Risk Yönetimi")
    
    risk_headers = [
        "🆔 Risk ID", "⚠️ Risk Adı", "📝 Açıklama", "📊 Olasılık", "💥 Etki", 
        "🎯 Risk Skoru", "👤 Sorumlu", "📅 Tarih", "🛡️ Önlem", "📊 Durum", "📋 Notlar"
    ]
    
    for col, header in enumerate(risk_headers, 1):
        cell = risk_ws.cell(row=1, column=col, value=header)
        cell.style = header_styles['header5']
    
    # Risk verileri
    risk_data = [
        ["R001", "🗄️ Veritabanı Performansı", "Büyük veri setlerinde performans düşüklüğü", "🟡 Orta", "🔴 Yüksek", "6", "Backend Dev", "15.09.2024", "İndexleme ve optimizasyon", "🟡 İzleniyor", "PostgreSQL optimizasyonu"],
        ["R002", "👤 Ekip Üyesi Ayrılması", "Kritik ekip üyesinin projeden ayrılması", "🟢 Düşük", "🔴 Yüksek", "3", "Proje Yöneticisi", "01.09.2024", "Yedek kaynak planı", "🟢 Kontrol Altında", "Backup developer hazır"],
        ["R003", "📅 Zaman Aşımı", "Belirtilen sürede tamamlanamama riski", "🟡 Orta", "🟡 Orta", "4", "Proje Yöneticisi", "01.09.2024", "Sprint planlaması", "🟡 İzleniyor", "Haftalık kontroller"],
        ["R004", "🔒 Güvenlik Açığı", "Sistemde güvenlik zafiyeti tespit edilmesi", "🟢 Düşük", "🔴 Yüksek", "3", "Security Expert", "20.09.2024", "Güvenlik testleri", "🟢 Kontrol Altında", "Penetrasyon testleri"],
        ["R005", "🌐 API Performansı", "Yüksek yükte API yanıt süresi artışı", "🟡 Orta", "🟡 Orta", "4", "Backend Dev", "25.09.2024", "Load balancing", "⚪ Yeni", "Yük testleri yapılacak"],
        ["R006", "📱 Mobil Uyumluluk", "Farklı cihazlarda görünüm problemleri", "🟡 Orta", "🟢 Düşük", "2", "Frontend Dev", "05.10.2024", "Responsive testing", "⚪ Yeni", "Cross-browser testler"],
        ["R007", "☁️ Deployment Hatası", "Canlı ortama geçişte teknik problemler", "🟡 Orta", "🔴 Yüksek", "6", "DevOps", "10.11.2024", "Staging ortamı", "⚪ Yeni", "Blue-green deployment"],
        ["R008", "💰 Bütçe Aşımı", "Planlanan bütçenin aşılması", "🟢 Düşük", "🟡 Orta", "2", "Proje Yöneticisi", "01.09.2024", "Bütçe takibi", "🟢 Kontrol Altında", "Haftalık bütçe raporları"]
    ]
    
    for row_num, risk in enumerate(risk_data, 2):
        for col_num, value in enumerate(risk, 1):
            cell = risk_ws.cell(row=row_num, column=col_num, value=value)
            
            # Risk skoru renklendirme
            if col_num == 6:  # Risk Skoru sütunu
                score = int(str(value)) if str(value).isdigit() else 0
                if score >= 6:
                    cell.fill = PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif score >= 4:
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                else:
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
            
            # Durum renklendirme
            elif col_num == 10:  # Durum sütunu
                if "Kontrol Altında" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "İzleniyor" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Yeni" in str(value):
                    cell.fill = PatternFill(start_color=colors['info'], end_color=colors['info'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
            
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Sütun genişlikleri
    column_widths_risk = [8, 20, 30, 12, 12, 10, 15, 12, 25, 15, 25]
    for i, width in enumerate(column_widths_risk, 1):
        risk_ws.column_dimensions[get_column_letter(i)].width = width
    
    # 6. BÜTÇE TAKİBİ SHEET
    budget_ws = wb.create_sheet("💰 Bütçe Takibi")
    
    budget_headers = [
        "🏷️ Kategori", "📋 Açıklama", "💰 Planlanan", "💸 Harcanan", 
        "💰 Kalan", "📊 Kullanım %", "📅 Tarih", "👤 Onaylayan", "📋 Notlar"
    ]
    
    for col, header in enumerate(budget_headers, 1):
        cell = budget_ws.cell(row=1, column=col, value=header)
        cell.style = header_styles['header1']
    
    # Bütçe verileri
    budget_data = [
        ["👥 İnsan Kaynağı", "Ekip üyelerinin maaşları", "80000₺", "15000₺", "65000₺", "19%", "27.08.2024", "Proje Yöneticisi", "3 aylık toplam"],
        ["🖥️ Teknoloji", "Sunucu, lisans, araçlar", "15000₺", "3000₺", "12000₺", "20%", "25.08.2024", "CTO", "AWS + lisanslar"],
        ["🎨 Tasarım", "UI/UX tasarım hizmetleri", "8000₺", "2000₺", "6000₺", "25%", "20.08.2024", "Tasarım Müdürü", "Figma Pro lisansı"],
        ["📚 Eğitim", "Ekip eğitimleri ve sertifikalar", "5000₺", "1000₺", "4000₺", "20%", "15.08.2024", "HR", "Django kursu"],
        ["🧪 Test", "QA araçları ve test hizmetleri", "7000₺", "500₺", "6500₺", "7%", "10.08.2024", "QA Lead", "Selenium Grid"],
        ["☁️ Deployment", "Hosting ve deployment maliyetleri", "10000₺", "1000₺", "9000₺", "10%", "05.08.2024", "DevOps", "Production sunucu"],
        ["📋 Dokümantasyon", "Teknik yazım ve dokümantasyon", "3000₺", "0₺", "3000₺", "0%", "01.08.2024", "Tech Writer", "Henüz başlanmadı"],
        ["🎯 Pazarlama", "Tanıtım ve pazarlama giderleri", "12000₺", "2000₺", "10000₺", "17%", "22.08.2024", "Pazarlama", "SEO optimizasyonu"],
        ["🔒 Güvenlik", "Güvenlik testleri ve sertifikalar", "8000₺", "0₺", "8000₺", "0%", "01.09.2024", "Security", "Penetrasyon testleri"],
        ["📊 Raporlama", "Proje takip araçları", "2000₺", "500₺", "1500₺", "25%", "18.08.2024", "PM", "Jira + Confluence"]
    ]
    
    for row_num, budget in enumerate(budget_data, 2):
        for col_num, value in enumerate(budget, 1):
            cell = budget_ws.cell(row=row_num, column=col_num, value=value)
            
            # Kullanım yüzdesi renklendirme
            if col_num == 6:  # Kullanım % sütunu
                percentage = str(value).replace('%', '')
                if percentage.isdigit():
                    pct = int(percentage)
                    if pct >= 80:
                        cell.fill = PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type="solid")
                        cell.font = Font(color='FFFFFF', bold=True)
                    elif pct >= 50:
                        cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                        cell.font = Font(color='FFFFFF', bold=True)
                    else:
                        cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                        cell.font = Font(color='FFFFFF', bold=True)
            
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Sütun genişlikleri
    column_widths_budget = [18, 25, 12, 12, 12, 12, 12, 15, 25]
    for i, width in enumerate(column_widths_budget, 1):
        budget_ws.column_dimensions[get_column_letter(i)].width = width
    
    # Toplam satırı ekle
    total_row = len(budget_data) + 3
    budget_ws.cell(row=total_row, column=1, value="📊 TOPLAM").font = Font(bold=True, size=12)
    budget_ws.cell(row=total_row, column=3, value="150000₺").font = Font(bold=True, size=12, color=colors['primary'])
    budget_ws.cell(row=total_row, column=4, value="25000₺").font = Font(bold=True, size=12, color=colors['danger'])
    budget_ws.cell(row=total_row, column=5, value="125000₺").font = Font(bold=True, size=12, color=colors['success'])
    budget_ws.cell(row=total_row, column=6, value="17%").font = Font(bold=True, size=12)
    
    # 7. İLETİŞİM VE NOTLAR SHEET
    communication_ws = wb.create_sheet("📞 İletişim & Notlar")
    
    comm_headers = [
        "📅 Tarih", "🎯 Toplantı Tipi", "👥 Katılımcılar", "📋 Konu", 
        "📝 Kararlar", "✅ Aksiyon", "👤 Sorumlu", "📅 Deadline", "📊 Durum", "📋 Notlar"
    ]
    
    for col, header in enumerate(comm_headers, 1):
        cell = communication_ws.cell(row=1, column=col, value=header)
        cell.style = header_styles['header2']
    
    # İletişim verileri
    comm_data = [
        ["25.08.2024", "🎯 Proje Kickoff", "Tüm Ekip", "Proje başlangıç toplantısı", "Proje scope belirlendi", "Ekip rolleri atanacak", "PM", "27.08.2024", "✅ Tamamlandı", "Motivasyonel toplantı"],
        ["26.08.2024", "📊 Gereksinim", "PM, Analyst, Müşteri", "Detaylı gereksinim analizi", "Fonksiyonel gereksinimler onaylandı", "Wireframe hazırlanacak", "Designer", "30.08.2024", "🟡 Devam Ediyor", "Müşteri çok detaycı"],
        ["27.08.2024", "⚙️ Teknik", "Backend, DevOps", "Altyapı mimarisi", "PostgreSQL + Django kararı", "Sunucu kurulacak", "DevOps", "02.09.2024", "⚪ Bekliyor", "AWS üzerinde olacak"],
        ["28.08.2024", "🎨 Tasarım", "Designer, Frontend", "UI/UX konsepti", "Material Design tercih edildi", "Figma prototipi", "Designer", "05.09.2024", "⚪ Bekliyor", "Mobil first yaklaşım"],
        ["01.09.2024", "📈 Haftalık", "Tüm Ekip", "İlerleme raporu", "Planlama %60 tamamlandı", "Tasarım başlatılacak", "Tüm Ekip", "08.09.2024", "⚪ Bekliyor", "Düzenli toplantı"],
        ["15.09.2024", "⚙️ Code Review", "Backend Team", "Model yapısı gözden geçirme", "Model tasarımı onaylandı", "Development başlatılacak", "Backend Lead", "18.09.2024", "⚪ Bekliyor", ""],
        ["30.09.2024", "🧪 Test Planlama", "QA, Backend, Frontend", "Test stratejisi", "Otomatik test kararı", "Test framework seçilecek", "QA Lead", "05.10.2024", "⚪ Bekliyor", "Pytest + Jest kullanılacak"],
        ["15.10.2024", "🚀 Deployment", "DevOps, Backend", "Canlıya çıkış planı", "Blue-green deployment", "Pipeline kurulacak", "DevOps", "20.10.2024", "⚪ Bekliyor", "GitHub Actions"],
        ["01.11.2024", "📊 Final Review", "Tüm Ekip, Müşteri", "Proje teslim öncesi", "Son kontroller", "Bug fixes", "Tüm Ekip", "10.11.2024", "⚪ Bekliyor", ""],
        ["20.11.2024", "🎉 Proje Teslimi", "Tüm Ekip, Müşteri", "Projenin resmi teslimi", "Müşteri kabulü", "Eğitim verilecek", "PM", "25.11.2024", "⚪ Bekliyor", "Kutlama!"]
    ]
    
    for row_num, comm in enumerate(comm_data, 2):
        for col_num, value in enumerate(comm, 1):
            cell = communication_ws.cell(row=row_num, column=col_num, value=value)
            
            # Durum renklendirme
            if col_num == 9:  # Durum sütunu
                if "Tamamlandı" in str(value):
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Devam" in str(value):
                    cell.fill = PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type="solid")
                    cell.font = Font(color='FFFFFF', bold=True)
                elif "Bekliyor" in str(value):
                    cell.fill = PatternFill(start_color=colors['light'], end_color=colors['light'], fill_type="solid")
            
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Sütun genişlikleri
    column_widths_comm = [12, 15, 20, 25, 30, 20, 15, 12, 15, 25]
    for i, width in enumerate(column_widths_comm, 1):
        communication_ws.column_dimensions[get_column_letter(i)].width = width
    
    # Dosyayı kaydet
    file_path = r"c:\Users\ahmet.yildirir\Desktop\Apps\ProjectsDoc\Django_Proje_Yonetimi_Taslagi.xlsx"
    wb.save(file_path)
    print(f"Excel dosyası oluşturuldu: {file_path}")
    
    return file_path

if __name__ == "__main__":
    create_django_project_management_excel()
