
import pandas as pd
from openpyxl import Workbook
import json

wb = Workbook()

default_sheet = wb.active
if default_sheet:
    wb.remove(default_sheet)

# JSON formatında örnek proje verileri
proje_json = '''
{
    "projeler": [
        {
            "proje_ad": "E-Ticaret Sitesi",
            "proje_aciklama": "Django ile geliştirilmiş online alışveriş platformu",
            "baslangic_tarihi": "2023-09-01",
            "bitis_tarihi": "2023-12-15",
            "ekip_uyeleri": ["Ahmet", "Mehmet", "Ayşe"],
            "teknolojiler": ["Django", "PostgreSQL", "HTML/CSS", "JavaScript"],
            "butce": 150000
        },
        {
            "proje_ad": "Mobil Uygulama",
            "proje_aciklama": "Flutter ile geliştirilmiş mobil alışveriş uygulaması",
            "baslangic_tarihi": "2023-10-01",
            "bitis_tarihi": "2024-01-31",
            "ekip_uyeleri": ["Ali", "Zeynep", "Can"],
            "teknolojiler": ["Flutter", "Firebase", "Dart"],
            "butce": 200000
        },
        {
            "proje_ad": "CRM Sistemi",
            "proje_aciklama": "Müşteri ilişkileri yönetim sistemi",
            "baslangic_tarihi": "2023-11-15",
            "bitis_tarihi": "2024-03-20",
            "ekip_uyeleri": ["Burak", "Selin", "Okan", "Deniz"],
            "teknolojiler": ["Django", "React", "MySQL", "Redis"],
            "butce": 300000
        }
    ]
}
'''

# JSON verilerini Python sözlüğüne dönüştürme
proje_verileri_json = json.loads(proje_json)

# Temel proje özelliklerini içeren sayfayı oluşturma
proje_sayfa = wb.create_sheet(title="Proje Özellikleri")
proje_sayfa_basliklar = ["Proje Ad", "Proje Açıklaması", "Başlangıç Tarihi", "Bitiş Tarihi", "Bütçe"]
proje_sayfa.append(proje_sayfa_basliklar)

# JSON'dan alınan verileri Excel'e ekleme
for proje in proje_verileri_json["projeler"]:
    proje_satir = [
        proje["proje_ad"],
        proje["proje_aciklama"],
        proje["baslangic_tarihi"],
        proje["bitis_tarihi"],
        proje["butce"]
    ]
    proje_sayfa.append(proje_satir)

# Ekip üyeleri için yeni bir sayfa oluşturma
ekip_sayfa = wb.create_sheet(title="Ekip Üyeleri")
ekip_sayfa.append(["Proje Ad", "Ekip Üyeleri"])

# Her projenin ekip üyelerini ekleme
for proje in proje_verileri_json["projeler"]:
    for uye in proje["ekip_uyeleri"]:
        ekip_sayfa.append([proje["proje_ad"], uye])

# Teknolojiler için yeni bir sayfa oluşturma
teknoloji_sayfa = wb.create_sheet(title="Teknolojiler")
teknoloji_sayfa.append(["Proje Ad", "Teknoloji"])

# Her projenin kullandığı teknolojileri ekleme
for proje in proje_verileri_json["projeler"]:
    for teknoloji in proje["teknolojiler"]:
        teknoloji_sayfa.append([proje["proje_ad"], teknoloji])

# JSON verilerini Excel dosyasında ayrı bir sayfa olarak sakla
json_sayfa = wb.create_sheet(title="JSON Veri")
json_sayfa.append(["Proje Verileri (JSON)"])
json_sayfa.append([proje_json])  # JSON verilerini bir hücreye yerleştir

# Excel dosyasını kaydet
file_path = r"C:\Users\ahmet.yildirir\Desktop\Apps\ProjectsDoc\TestExcel.xlsx"
wb.save(file_path)
print(f"Excel dosyası oluşturuldu: {file_path}")
print("JSON veriler dahil edildi ve ekstra sayfalar eklendi.")