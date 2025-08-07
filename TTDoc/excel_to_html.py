import pandas as pd
import openpyxl
from pathlib import Path
import json

def analyze_excel_file(file_path):
    """Excel dosyasını analiz et ve tüm sekmeler hakkında bilgi topla"""
    
    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    # Tüm sekme isimleri
    sheet_names = workbook.sheetnames
    print(f"Toplam {len(sheet_names)} sekme bulundu:")
    for i, name in enumerate(sheet_names, 1):
        print(f"{i}. {name}")
    
    # Her sekme için detaylı analiz
    analysis_results = {}
    
    for sheet_name in sheet_names:
        print(f"\n--- {sheet_name} sekmesi analiz ediliyor ---")
        
        # Pandas ile sekmeyi oku
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Sekme hakkında temel bilgiler
            sheet_info = {
                'name': sheet_name,
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns),
                'has_data': not df.empty,
                'sample_data': df.head(10).to_dict('records') if not df.empty else [],
                'data_types': df.dtypes.to_dict() if not df.empty else {},
                'non_null_counts': df.count().to_dict() if not df.empty else {}
            }
            
            # Openpyxl ile ek bilgiler al
            worksheet = workbook[sheet_name]
            sheet_info['max_row'] = worksheet.max_row
            sheet_info['max_column'] = worksheet.max_column
            
            # Boş olmayan hücre sayısı
            non_empty_cells = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        non_empty_cells += 1
            sheet_info['non_empty_cells'] = non_empty_cells
            
            analysis_results[sheet_name] = sheet_info
            
            print(f"  - Satır sayısı: {sheet_info['rows']}")
            print(f"  - Sütun sayısı: {sheet_info['columns']}")
            print(f"  - Boş olmayan hücre sayısı: {non_empty_cells}")
            print(f"  - Sütun isimleri: {sheet_info['column_names'][:5]}{'...' if len(sheet_info['column_names']) > 5 else ''}")
            
        except Exception as e:
            print(f"  - Hata: {str(e)}")
            analysis_results[sheet_name] = {
                'name': sheet_name,
                'error': str(e),
                'has_data': False
            }
    
    return analysis_results

if __name__ == "__main__":
    file_path = "Ağustos_TT.xlsx"
    
    if Path(file_path).exists():
        results = analyze_excel_file(file_path)
        
        # Sonuçları JSON dosyasına kaydet
        with open('excel_analysis.json', 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2, default=str)
        
        print(f"\nAnaliz tamamlandı. Sonuçlar 'excel_analysis.json' dosyasına kaydedildi.")
    else:
        print(f"Dosya bulunamadı: {file_path}")
