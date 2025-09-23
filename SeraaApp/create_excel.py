#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_sera_excel():
    # Excel dosyası oluştur
    wb = Workbook()
    
    # Stil tanımlamaları
    header_font = Font(bold=True, color='000000')
    header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ============================================
    # 1. SEKME - PROJE TANIMLARI
    # ============================================
    ws1 = wb.active
    ws1.title = '1-Proje Tanımları'
    
    # Proje bilgileri
    proje_data = [
        ['PROJE BİLGİLERİ', 'DEĞER', 'BİRİM'],
        ['Tünel Uzunluğu', 250, 'm'],
        ['Tünel Sayısı', 50, 'adet'],
        ['Tünel Genişlik', 9.6, 'm'],
        ['Duvar Kolon Arası', 2.5, 'm'],
        ['Orta Kolon Aralığı', 5, 'm'],
        ['Oluk Boyu', 5, 'm'],
        ['Tek Akıntı', 1, 'adet'],
        ['Çift Akıntılı', 2, 'adet'],
        ['Duvar Üstü', 2, 'adet'],
        ['Baştaki Kolonlar', 102, 'adet'],
        ['', '', ''],
        ['HESAPLANAN DEĞERLER', 'SONUÇ', 'BİRİM'],
        ['Toplam Alan', '=B2*B3*B4', 'm²'],
        ['Toplam Montaj Noktası', '=(B2/B5+1)*(B3+1)', 'adet'],
        ['Toplam Direk Uzunluğu', '=B2*B3', 'm']
    ]
    
    # Verileri sayfaya yaz
    for row_idx, row in enumerate(proje_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx in [1, 13]:  # Başlık satırları
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sütun genişlikleri
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 10
    
    # ============================================
    # 2. SEKME - MONTAJ NOKTASI MALİYETİ
    # ============================================
    ws2 = wb.create_sheet('2-Montaj Noktası')
    
    montaj_data = [
        ['NO', 'İMAL KODU', 'MALZEME TANIMI', 'MİKTAR', 'BİRİM FİYAT', 'BİRİM AĞIRLIK', 'TOPLAM MALİYET'],
        [1, 'BP-01', '70*70*1200*2,00 mm Ankaraj', 1, 125.50, 5.1, '=D2*E2'],
        [2, 'BP-02', '80*80*3000*3,00 mm Direk', 1, 285.75, 12.8, '=D3*E3'],
        [3, 'BP-03', 'Montaj Plakası 200*200*8mm', 2, 45.25, 2.5, '=D4*E4'],
        [4, 'BP-04', 'Bağlantı Elemanı M12*80', 4, 8.75, 0.2, '=D5*E5'],
        [5, 'BP-05', 'Kaynak İşçiliği', 1, 65.00, '-', '=D6*E6'],
        ['', '', '', '', '', '', ''],
        ['', '1 MONTAJ NOKTASI TOPLAM MALİYET', '', '', '', '', '=SUM(G2:G6)']
    ]
    
    # Verileri sayfaya yaz
    for row_idx, row in enumerate(montaj_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1 or row_idx == 8:  # Başlık satırları
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sütun genişlikleri
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 8
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 15
    
    # ============================================
    # 3. SEKME - 1 METRE DİREK MALİYETİ
    # ============================================
    ws3 = wb.create_sheet('3-Metre Direk')
    
    direk_data = [
        ['NO', 'İMAL KODU', 'MALZEME TANIMI', 'MİKTAR', 'BİRİM FİYAT', 'BİRİM AĞIRLIK', 'TOPLAM MALİYET'],
        [1, 'MD-01', '80*80*1000*2,5mm Ara Direk', 1, 95.25, 7.8, '=D2*E2'],
        [2, 'MD-02', 'Bağlantı Flanşı', 2, 25.50, 1.2, '=D3*E3'],
        [3, 'MD-03', 'Galvaniz Boyama', 1, 15.75, '-', '=D4*E4'],
        ['', '', '', '', '', '', ''],
        ['', '1 METRE DİREK TOPLAM MALİYET', '', '', '', '', '=SUM(G2:G4)']
    ]
    
    # Verileri sayfaya yaz
    for row_idx, row in enumerate(direk_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1 or row_idx == 6:  # Başlık satırları
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sütun genişlikleri
    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 8
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 15
    
    # ============================================
    # 4. SEKME - MÜŞTEMİLAT DİREK MALİYETİ
    # ============================================
    ws4 = wb.create_sheet('4-Müştemilat Direk')
    
    mustemilat_data = [
        ['NO', 'İMAL KODU', 'MALZEME TANIMI', 'MİKTAR', 'BİRİM FİYAT', 'BİRİM AĞIRLIK', 'TOPLAM MALİYET'],
        [1, 'MS-01', '80*80*2500*3,0mm Müştemilat Direk', 1, 238.50, 18.5, '=D2*E2'],
        [2, 'MS-02', 'Temel Ankrajı 70*70*800mm', 1, 85.25, 3.8, '=D3*E3'],
        [3, 'MS-03', 'Üst Bağlantı Aparatı', 1, 45.75, 2.2, '=D4*E4'],
        [4, 'MS-04', 'Montaj İşçiliği', 1, 75.00, '-', '=D5*E5'],
        ['', '', '', '', '', '', ''],
        ['', '1 MÜŞTEMİLAT DİREK TOPLAM MALİYET', '', '', '', '', '=SUM(G2:G5)']
    ]
    
    # Verileri sayfaya yaz
    for row_idx, row in enumerate(mustemilat_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1 or row_idx == 7:  # Başlık satırları
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sütun genişlikleri
    ws4.column_dimensions['A'].width = 5
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 30
    ws4.column_dimensions['D'].width = 8
    ws4.column_dimensions['E'].width = 12
    ws4.column_dimensions['F'].width = 12
    ws4.column_dimensions['G'].width = 15
    
    # ============================================
    # 5. SEKME - TOPLAM MALİYET HESAPLAMA
    # ============================================
    ws5 = wb.create_sheet('5-Toplam Maliyet')
    
    maliyet_data = [
        ['SERA PROJESİ TOPLAM MALİYET HESAPLAMA', '', '', ''],
        ['', '', '', ''],
        ['MALİYET KALEMİ', 'MİKTAR', 'BİRİM MALİYET', 'TOPLAM MALİYET'],
        ['Montaj Noktası Sayısı', "='1-Proje Tanımları'.B15", "='2-Montaj Noktası'.G8", '=B4*C4'],
        ['Ara Direk Uzunluğu (metre)', "='1-Proje Tanımları'.B16", "='3-Metre Direk'.G6", '=B5*C5'],
        ['Müştemilat Direk Sayısı', 25, "='4-Müştemilat Direk'.G7", '=B6*C6'],
        ['', '', '', ''],
        ['TOPLAM MALZEME MALİYETİ', '', '', '=SUM(D4:D6)'],
        ['İşçilik %15', '', '', '=D8*0.15'],
        ['Nakliye %5', '', '', '=D8*0.05'],
        ['Kar %10', '', '', '=D8*0.10'],
        ['', '', '', ''],
        ['TOPLAM PROJE MALİYETİ (KDV HARİÇ)', '', '', '=SUM(D8:D11)'],
        ['KDV %20', '', '', '=D13*0.20'],
        ['TOPLAM PROJE MALİYETİ (KDV DAHİL)', '', '', '=D13+D14'],
        ['', '', '', ''],
        ['M² BAŞINA MALİYET HESAPLAMA', '', '', ''],
        ['Toplam Alan (m²)', "='1-Proje Tanımları'.B14", '', ''],
        ['M² Başına Maliyet (KDV Hariç)', '', '', '=D13/B18'],
        ['M² Başına Maliyet (KDV Dahil)', '', '', '=D15/B18']
    ]
    
    # Verileri sayfaya yaz
    for row_idx, row in enumerate(maliyet_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx in [1, 3, 8, 13, 15, 17]:  # Başlık satırları
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # İlk satırı birleştir
    ws5.merge_cells('A1:D1')
    
    # Sütun genişlikleri
    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 15
    ws5.column_dimensions['C'].width = 15
    ws5.column_dimensions['D'].width = 18
    
    # Dosyayı kaydet
    filename = 'Sera_Proje_Maliyet_Hesaplama.xlsx'
    wb.save(filename)
    print(f"Excel dosyası başarıyla oluşturuldu: {filename}")
    
    return filename

if __name__ == "__main__":
    create_sera_excel()
