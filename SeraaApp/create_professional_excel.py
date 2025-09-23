#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle

def create_professional_sera_excel():
    # Excel dosyası oluştur
    wb = Workbook()
    
    # Profesyonel stil tanımlamaları
    title_font = Font(bold=True, size=16, color='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    sub_header_font = Font(bold=True, color='1F4E79', size=11)
    
    # Renk paleti
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')  # Koyu mavi
    input_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')   # SARI - INPUT
    calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')    # Açık yeşil
    result_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # Açık mavi
    warning_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Turuncu
    
    # Kenarlık stilleri
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), 
                         top=Side(style='medium'), bottom=Side(style='medium'))
    
    # ============================================
    # 1. SEKME - PROJE YÖNETİMİ
    # ============================================
    ws1 = wb.active
    ws1.title = '1-Proje Yönetimi'
    
    # Başlık
    ws1['B2'] = 'SERA PROJESİ YÖNETİM PANELİ'
    ws1['B2'].font = title_font
    ws1.merge_cells('B2:G2')
    
    # Proje bilgileri formu
    project_section = [
        ['', '', '', '', '', '', ''],
        ['PROJE BİLGİLERİ', '', '', '', 'PROJE DURUMU', '', ''],
        ['', '', '', '', '', '', ''],
        ['Proje Adı:', '', '', '', 'Hazırlama Tarihi:', '', '=TODAY()'],
        ['Proje Kodu:', '', 'SRA-2025-001', '', 'Son Güncelleme:', '', '=NOW()'],
        ['Müşteri Firma:', '', '', '', 'Durum:', '', 'Hesaplama Aşamasında'],
        ['İletişim Kişisi:', '', '', '', 'Onay Durumu:', '', 'Müşteri Onayı Bekliyor'],
        ['Telefon:', '', '', '', 'Revizyon:', '', '1.0'],
        ['E-posta:', '', '', '', '', '', ''],
        ['Proje Adresi:', '', '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        
        # Proje özeti
        ['PROJE ÖZETİ', '', '', '', 'MALİYET ÖZETİ', '', ''],
        ['', '', '', '', '', '', ''],
        ['Toplam Alan:', '', "='5-Maliyet Hesaplama'.C11", 'm²', 'Toplam Maliyet:', '', "='5-Maliyet Hesaplama'.C35"],
        ['Montaj Noktası:', '', "='5-Maliyet Hesaplama'.C12", 'adet', 'M² Başına:', '', "='5-Maliyet Hesaplama'.C37"],
        ['Ara Direk:', '', "='5-Maliyet Hesaplama'.C13", 'm', 'KDV Dahil:', '', "='5-Maliyet Hesaplama'.C36"],
        ['Müştemilat:', '', "='5-Maliyet Hesaplama'.C14", 'adet', 'Kar Oranı:', '', '%12'],
        ['', '', '', '', '', '', ''],
        
        # Notlar ve uyarılar
        ['NOTLAR VE UYARILAR', '', '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['• Sarı hücreler veri giriş alanlarıdır', '', '', '', '', '', ''],
        ['• Maliyet hesaplama 5. sekmede yapılır', '', '', '', '', '', ''],
        ['• Malzeme fiyatları 2-3-4. sekmelerde', '', '', '', '', '', ''],
        ['• Değişiklik yapmadan önce dosyayı kaydedin', '', '', '', '', '', '']
    ]
    
    for row_idx, row in enumerate(project_section, 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Başlık bölümleri
            if value and ('BİLGİLERİ' in str(value) or 'DURUMU' in str(value) or 'ÖZETİ' in str(value) or 'UYARILAR' in str(value)):
                cell.font = sub_header_font
                cell.fill = header_fill
                cell.font = Font(bold=True, color='FFFFFF', size=11)
            
            # Input alanları - SARI
            elif col_idx == 3 and row_idx in [6, 7, 8, 9, 10, 11, 12]:
                cell.fill = input_fill
            
            # Hesaplanan değerler
            elif '=' in str(value):
                cell.fill = calc_fill
            
            # Uyarı mesajları
            elif '•' in str(value):
                cell.fill = warning_fill
                cell.font = Font(italic=True, color='D35400')
    
    # Sütun genişlikleri
    column_widths = [3, 18, 20, 8, 18, 5, 20]
    for idx, width in enumerate(column_widths, 1):
        ws1.column_dimensions[get_column_letter(idx)].width = width
    
    # ============================================
    # 2. SEKME - MONTAJ MALZEMELER (Gelişmiş)
    # ============================================
    ws2 = wb.create_sheet('2-Montaj Malzemeleri')
    
    # Başlık ve açıklama
    ws2['A1'] = '🔧 MONTAJ NOKTASI MALZEME LİSTESİ (1 ADET İÇİN)'
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:J1')
    
    ws2['A2'] = '💡 Bu sayfadaki maliyetler 1 montaj noktası içindir. Toplam hesaplama ana sayfada yapılır.'
    ws2['A2'].font = Font(italic=True, color='7030A0', size=10)
    ws2.merge_cells('A2:J2')
    
    # Tablo başlıkları
    headers = ['S.No', 'Malzeme Kodu', 'Malzeme Tanımı', 'Miktar', 'Birim', 'Birim Fiyat (₺)', 'Toplam (₺)', 'Tedarikçi', 'Son Güncelleme', 'Açıklama']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Detaylı malzeme listesi
    montaj_materials = [
        [1, 'MT-001', '70x70x1200x2.0mm Galvanizli Çelik Ankaraj', 1, 'adet', 145.50, '=D5*F5', 'Demir A.Ş.', '=TODAY()', 'Temel ankraj sistemi'],
        [2, 'MT-002', '80x80x3000x3.0mm Yapısal Çelik Direk', 1, 'adet', 325.75, '=D6*F6', 'Çelik Ltd.', '=TODAY()', 'Ana taşıyıcı direk'],
        [3, 'MT-003', '200x200x10mm Montaj Plaka Takımı', 2, 'adet', 65.25, '=D7*F7', 'Metal Co.', '=TODAY()', 'Bağlantı plakası'],
        [4, 'MT-004', 'M12x80 Galvanizli Cıvata Seti (4lü)', 1, 'takım', 28.75, '=D8*F8', 'Bağlantı Ltd.', '=TODAY()', 'Montaj cıvataları'],
        [5, 'MT-005', 'Kaynak İşçiliği (Kalifiye)', 2, 'saat', 75.00, '=D9*F9', 'Kaynak Team', '=TODAY()', 'Profesyonel kaynak'],
        [6, 'MT-006', 'Galvaniz Sıcak Daldırma', 1, 'kg', 8.50, '=D10*F10', 'Galvaniz A.Ş.', '=TODAY()', 'Korozyon koruması'],
        [7, 'MT-007', 'L50x50x5 Destek Profili', 1.5, 'metre', 42.25, '=D11*F11', 'Profil Ltd.', '=TODAY()', 'Ek destek elemanı'],
        [8, 'MT-008', 'Elastomerik Conta Sistemi', 1, 'takım', 15.75, '=D12*F12', 'Conta Co.', '=TODAY()', 'Hava geçirmezlik'],
        ['', '', '', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', '', '', ''],
        ['', 'TOPLAM MONTAJ MALİYETİ (1 ADET)', '', '', '', '', '=SUM(G5:G50)', '', '', '']
    ]
    
    for row_idx, row in enumerate(montaj_materials, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Input alanları - SARI (Miktar ve Fiyat)
            if col_idx in [4, 6] and row_idx <= 12:
                cell.fill = input_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Hesaplanan değerler - YEŞİL
            elif col_idx == 7 and '=' in str(value):
                cell.fill = calc_fill
                cell.alignment = Alignment(horizontal='right')
            
            # Toplam satırı
            elif row_idx == 15 and col_idx == 7:
                cell.font = Font(bold=True, size=12, color='1F4E79')
                cell.fill = result_fill
                cell.border = thick_border
    
    # Sütun genişlikleri
    column_widths = [5, 12, 35, 8, 8, 12, 12, 15, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validation - Birim seçimi
    dv_unit = DataValidation(type="list", formula1='"adet,metre,kg,takım,saat,m2,m3,ton"')
    ws2.add_data_validation(dv_unit)
    dv_unit.add('E5:E50')
    
    # ============================================
    # 3. SEKME - DİREK MALZEMELER (Gelişmiş)
    # ============================================
    ws3 = wb.create_sheet('3-Direk Malzemeleri')
    
    # Başlık
    ws3['A1'] = '📏 ARA DİREK MALZEME LİSTESİ (1 METRE İÇİN)'
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:J1')
    
    ws3['A2'] = '💡 Bu sayfadaki maliyetler 1 metre direk için hesaplanmıştır. Toplam metreyle çarpılacak.'
    ws3['A2'].font = Font(italic=True, color='7030A0', size=10)
    ws3.merge_cells('A2:J2')
    
    # Başlıklar
    for col_idx, header in enumerate(headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Direk malzemeleri
    direk_materials = [
        [1, 'DR-001', '80x80x1000x2.5mm Ara Direk Profili', 1, 'metre', 115.25, '=D5*F5', 'Profil A.Ş.', '=TODAY()', 'Yapısal ara direk'],
        [2, 'DR-002', '120x8mm Flanş Bağlantı Sistemi', 2, 'adet', 35.50, '=D6*F6', 'Bağlantı Ltd.', '=TODAY()', 'Direk bağlantı flanşı'],
        [3, 'DR-003', 'Elektrostatik Toz Boya', 0.8, 'm2', 25.75, '=D7*F7', 'Boya Co.', '=TODAY()', 'Koruyucu boya sistemi'],
        [4, 'DR-004', 'U Tipi Montaj Kelepçesi', 2, 'adet', 18.25, '=D8*F8', 'Kelepçe Ltd.', '=TODAY()', 'Hızlı montaj sistemi'],
        [5, 'DR-005', 'EPDM Conta ve Sızdırmazlık', 1, 'takım', 12.50, '=D9*F9', 'Conta A.Ş.', '=TODAY()', 'Hava geçirmezlik'],
        [6, 'DR-006', 'Darbe Emici Ped', 1, 'adet', 8.75, '=D10*F10', 'Ped Co.', '=TODAY()', 'Titreşim önleyici'],
        ['', '', '', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', '', '', ''],
        ['', 'TOPLAM DİREK MALİYETİ (1 METRE)', '', '', '', '', '=SUM(G5:G50)', '', '', '']
    ]
    
    for row_idx, row in enumerate(direk_materials, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [4, 6] and row_idx <= 10:
                cell.fill = input_fill
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 7 and '=' in str(value):
                cell.fill = calc_fill
                cell.alignment = Alignment(horizontal='right')
            elif row_idx == 13 and col_idx == 7:
                cell.font = Font(bold=True, size=12, color='1F4E79')
                cell.fill = result_fill
                cell.border = thick_border
    
    # Sütun genişlikleri
    for idx, width in enumerate(column_widths, 1):
        ws3.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validation
    dv_unit2 = DataValidation(type="list", formula1='"adet,metre,kg,takım,saat,m2,m3,ton"')
    ws3.add_data_validation(dv_unit2)
    dv_unit2.add('E5:E50')
    
    # ============================================
    # 4. SEKME - MÜŞTEMİLAT MALZEMELER (Gelişmiş)
    # ============================================
    ws4 = wb.create_sheet('4-Müştemilat Malzemeleri')
    
    # Başlık
    ws4['A1'] = '🏢 MÜŞTEMİLAT DİREK MALZEME LİSTESİ (1 DİREK İÇİN)'
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:J1')
    
    ws4['A2'] = '💡 Müştemilat alanı (teknik oda/kazan dairesi) direkleri için özel malzeme listesi.'
    ws4['A2'].font = Font(italic=True, color='7030A0', size=10)
    ws4.merge_cells('A2:J2')
    
    # Başlıklar
    for col_idx, header in enumerate(headers, 1):
        cell = ws4.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Müştemilat malzemeleri
    mustemilat_materials = [
        [1, 'MS-001', '80x80x2500x3.0mm Ağır Hizmet Direği', 1, 'adet', 285.50, '=D5*F5', 'Ağır Çelik Ltd.', '=TODAY()', 'Yüksek kapasiteli direk'],
        [2, 'MS-002', '70x70x800mm Özel Temel Ankrajı', 1, 'adet', 125.25, '=D6*F6', 'Temel A.Ş.', '=TODAY()', 'Güçlendirilmiş ankraj'],
        [3, 'MS-003', 'Ağır Yük Taşıma Başlığı', 1, 'adet', 85.75, '=D7*F7', 'Başlık Co.', '=TODAY()', 'Tavan bağlantı sistemi'],
        [4, 'MS-004', 'Özel Montaj İşçiliği (Kalifiye)', 3, 'saat', 95.00, '=D8*F8', 'Uzman Team', '=TODAY()', 'Teknik montaj hizmeti'],
        [5, 'MS-005', 'C25 Beton Temeli (0.6m³)', 0.6, 'm3', 220.00, '=D9*F9', 'Beton A.Ş.', '=TODAY()', 'Yüksek dayanımlı beton'],
        [6, 'MS-006', 'Φ14 Nervürlü Donatı', 25, 'kg', 12.50, '=D10*F10', 'Demir Ltd.', '=TODAY()', 'Temel donatı sistemi'],
        [7, 'MS-007', 'Su Yalıtım Membranı', 1.2, 'm2', 45.00, '=D11*F11', 'Yalıtım Co.', '=TODAY()', 'Nem ve su koruması'],
        [8, 'MS-008', 'Titreşim Önleyici Pad', 1, 'takım', 35.75, '=D12*F12', 'Titreşim Ltd.', '=TODAY()', 'Makine titreşim yalıtımı'],
        ['', '', '', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', '', '', ''],
        ['', 'TOPLAM MÜŞTEMİLAT MALİYETİ (1 DİREK)', '', '', '', '', '=SUM(G5:G50)', '', '', '']
    ]
    
    for row_idx, row in enumerate(mustemilat_materials, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [4, 6] and row_idx <= 12:
                cell.fill = input_fill
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 7 and '=' in str(value):
                cell.fill = calc_fill
                cell.alignment = Alignment(horizontal='right')
            elif row_idx == 15 and col_idx == 7:
                cell.font = Font(bold=True, size=12, color='1F4E79')
                cell.fill = result_fill
                cell.border = thick_border
    
    # Sütun genişlikleri
    for idx, width in enumerate(column_widths, 1):
        ws4.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validation
    dv_unit3 = DataValidation(type="list", formula1='"adet,metre,kg,takım,saat,m2,m3,ton"')
    ws4.add_data_validation(dv_unit3)
    dv_unit3.add('E5:E50')
    
    # ============================================
    # 5. SEKME - PROFESYONEL MALİYET HESAPLAMA
    # ============================================
    ws5 = wb.create_sheet('5-Maliyet Hesaplama')
    
    # Ana başlık
    ws5['A1'] = '💰 SERA PROJESİ PROFESYONEL MALİYET HESAPLAMA VE ANALİZ MERKEZİ'
    ws5['A1'].font = Font(bold=True, size=16, color='1F4E79')
    ws5.merge_cells('A1:H1')
    
    # Ana hesaplama bölümü
    main_calc_section = [
        ['', '', '', '', '', '', '', ''],
        ['📊 PROJE PARAMETRELERİ VE HESAPLAMA', '', '', '', '📈 OTOMATIK HESAPLANAN DEĞERLER', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['Parametre', '🟡 Değer', 'Birim', 'Açıklama', 'Hesaplama', 'Sonuç', 'Birim', 'Durum'],
        
        # Sol taraf - INPUT parametreleri (SARI)
        ['Tünel Uzunluğu', 250, 'm', 'Ana sera uzunluğu', 'Toplam Alan', '=B5*B6*B7', 'm²', '✓'],
        ['Tünel Sayısı', 50, 'adet', 'Paralel tünel adedi', 'Montaj Noktası', '=(B5/B9+1)*(B6+1)', 'adet', '✓'],
        ['Tünel Genişlik', 9.6, 'm', 'Her tünel genişliği', 'Ara Direk Uzunluğu', '=B5*B6', 'm', '✓'],
        ['Duvar Kolon Arası', 2.5, 'm', 'Duvar direk aralığı', 'Müştemilat Alan', '=F5*0.08', 'm²', '✓'],
        ['Orta Kolon Aralığı', 5, 'm', 'İç direk aralığı', 'Müştemilat Direk', 'BELİRLENECEK', 'adet', '⚠'],
        ['', '', '', '', '', '', '', ''],
        
        # Manuel input alanları
        ['📝 MANUEL GİRİŞ ALANLARI', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['Montaj Noktası Sayısı', 'MANUEL', 'adet', '👈 Buraya sayı girin', '', '', '', ''],
        ['Ara Direk Uzunluğu', 'MANUEL', 'm', '👈 Buraya sayı girin', '', '', '', ''],
        ['Müştemilat Direk Sayısı', 'MANUEL', 'adet', '👈 Buraya sayı girin', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        
        # Birim maliyetler - diğer sayfalardan otomatik
        ['💸 BİRİM MALİYETLER (Diğer Sayfalardan Otomatik)', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['1 Montaj Noktası Maliyeti', "='2-Montaj Malzemeleri'.G15", '₺/adet', 'Otomatik hesaplanan', '', '', '', '✓'],
        ['1 Metre Direk Maliyeti', "='3-Direk Malzemeleri'.G13", '₺/m', 'Otomatik hesaplanan', '', '', '', '✓'],
        ['1 Müştemilat Direk Maliyeti', "='4-Müştemilat Malzemeleri'.G15", '₺/adet', 'Otomatik hesaplanan', '', '', '', '✓'],
        ['', '', '', '', '', '', '', ''],
        
        # Ana maliyet hesaplama
        ['🎯 ANA MALİYET HESAPLAMA', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['Montaj Toplam Maliyeti', '=B13*B18', '₺', 'Montaj × Birim', '', '', '', '✓'],
        ['Direk Toplam Maliyeti', '=B14*B19', '₺', 'Direk × Birim', '', '', '', '✓'],
        ['Müştemilat Toplam Maliyeti', '=B15*B20', '₺', 'Müştemilat × Birim', '', '', '', '✓'],
        ['', '', '', '', '', '', '', ''],
        ['🏗️ TOPLAM MALZEME MALİYETİ', '=SUM(B23:B25)', '₺', 'Temel maliyet', '', '', '', '✓'],
        ['', '', '', '', '', '', '', ''],
        
        # Ek maliyetler ve marjlar
        ['📋 EK MALİYETLER VE MARJLAR', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['İşçilik ve Montaj (%)', 15, '%', '👈 Oran değiştirilebilir', 'İşçilik Tutarı', '=B27*B30/100', '₺', '✓'],
        ['Nakliye ve Lojistik (%)', 5, '%', '👈 Oran değiştirilebilir', 'Nakliye Tutarı', '=B27*B31/100', '₺', '✓'],
        ['Proje Yönetimi (%)', 3, '%', '👈 Oran değiştirilebilir', 'Yönetim Tutarı', '=B27*B32/100', '₺', '✓'],
        ['Sigorta ve Risk (%)', 2, '%', '👈 Oran değiştirilebilir', 'Sigorta Tutarı', '=B27*B33/100', '₺', '✓'],
        ['Kar Marjı (%)', 12, '%', '👈 Oran değiştirilebilir', 'Kar Tutarı', '=B27*B34/100', '₺', '✓'],
        ['', '', '', '', '', '', '', ''],
        
        # Nihai sonuçlar
        ['🎯 NİHAİ SONUÇLAR', '', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['Toplam (KDV Hariç)', '=B27+SUM(F30:F34)', '₺', 'Ana toplam', '', '', '', '✅'],
        ['KDV (%20)', '=B38*0.20', '₺', 'Vergi', '', '', '', '✅'],
        ['GENEL TOPLAM (KDV Dahil)', '=B38+B39', '₺', 'Final fiyat', '', '', '', '✅'],
        ['', '', '', '', '', '', '', ''],
        ['📊 M² BAŞINA MALİYET ANALİZİ', '', '', '', '', '', '', ''],
        ['M² Başına (KDV Hariç)', '=B38/F5', '₺/m²', 'Birim maliyet', '', '', '', '📊'],
        ['M² Başına (KDV Dahil)', '=B40/F5', '₺/m²', 'Birim fiyat', '', '', '', '📊']
    ]
    
    for row_idx, row in enumerate(main_calc_section, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Ana başlık satırları
            if '📊' in str(value) or '📝' in str(value) or '💸' in str(value) or '🎯' in str(value) or '📋' in str(value):
                cell.font = Font(bold=True, size=12, color='1F4E79')
                cell.fill = header_fill
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                if col_idx == 1:
                    ws5.merge_cells(f'{get_column_letter(col_idx)}{row_idx}:{get_column_letter(col_idx+3)}{row_idx}')
            
            # Input alanları - SARI
            elif col_idx == 2 and (row_idx in [5, 6, 7, 8, 9, 13, 14, 15, 30, 31, 32, 33, 34] or 'MANUEL' in str(value)):
                cell.fill = input_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thick_border
            
            # Hesaplanan değerler - YEŞİL
            elif '=' in str(value) and col_idx in [2, 6]:
                cell.fill = calc_fill
                cell.alignment = Alignment(horizontal='right')
            
            # Önemli sonuçlar - MAVİ
            elif row_idx in [38, 40] and col_idx == 2:
                cell.font = Font(bold=True, size=14, color='1F4E79')
                cell.fill = result_fill
                cell.border = thick_border
    
    # Sütun genişlikleri
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 18
    ws5.column_dimensions['C'].width = 8
    ws5.column_dimensions['D'].width = 20
    ws5.column_dimensions['E'].width = 18
    ws5.column_dimensions['F'].width = 18
    ws5.column_dimensions['G'].width = 8
    ws5.column_dimensions['H'].width = 8
    
    # Dosyayı kaydet
    filename = 'Sera_Professional_Ultimate_System.xlsx'
    wb.save(filename)
    print(f"🚀 PROFESYONEL SERA SİSTEMİ oluşturuldu: {filename}")
    print("\n✨ PROFESYONEL ÖZELLİKLER:")
    print("🟡 SARI = INPUT alanları (net olarak belirtildi)")
    print("🔗 Sayfalar arası tam entegrasyon")
    print("📊 Gelişmiş hesaplama sistemi") 
    print("💎 Profesyonel tasarım")
    print("📱 Kullanıcı dostu arayüz")
    print("⚡ Etkileşimli hesaplama")
    print("🎯 Manuel input alanları")
    print("📈 Detaylı analiz ve raporlama")
    
    return filename

if __name__ == "__main__":
    create_professional_sera_excel()
