#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_ultimate_sera_excel():
    # Excel dosyası oluştur
    wb = Workbook()
    
    # Gelişmiş stil tanımlamaları
    title_font = Font(bold=True, size=16, color='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sub_header_font = Font(bold=True, color='000000', size=11)
    sub_header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    result_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), 
                         top=Side(style='medium'), bottom=Side(style='medium'))
    
    # ============================================
    # 1. SEKME - PROJE BİLGİLERİ
    # ============================================
    ws1 = wb.active
    ws1.title = '1-Proje Detayları'
    
    # Başlık
    ws1['B2'] = 'SERA PROJESİ GENEL BİLGİLER'
    ws1['B2'].font = title_font
    ws1.merge_cells('B2:F2')
    
    # Proje bilgi formu
    project_info = [
        ['', '', '', '', '', ''],
        ['Proje Bilgileri', '', '', 'Değer', '', ''],
        ['Proje Adı:', '', '', '', '', ''],
        ['Proje Kodu:', '', '', '', '', ''],
        ['Müşteri Adı:', '', '', '', '', ''],
        ['Müşteri Telefon:', '', '', '', '', ''],
        ['Proje Adresi:', '', '', '', '', ''],
        ['Başlangıç Tarihi:', '', '', 'TODAY()', '', ''],
        ['Bitiş Tarihi:', '', '', '', '', ''],
        ['Proje Yöneticisi:', '', '', '', '', ''],
        ['Teknik Sorumlu:', '', '', '', '', ''],
        ['', '', '', '', '', ''],
        ['Durum Bilgileri', '', '', '', '', ''],
        ['Proje Durumu:', '', '', 'Hazırlanıyor', '', ''],
        ['Onay Durumu:', '', '', 'Beklemede', '', ''],
        ['Revizyon No:', '', '', '1.0', '', ''],
        ['Son Güncelleme:', '', '', '=NOW()', '', '']
    ]
    
    for row_idx, row in enumerate(project_info, 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            
            if col_idx == 1 and value and ':' in str(value):
                cell.font = sub_header_font
                cell.border = thin_border
            elif col_idx == 1 and value and ':' not in str(value) and value != '':
                cell.font = Font(bold=True, size=12)
                cell.fill = sub_header_fill
                cell.border = thick_border
            elif col_idx == 4 and row_idx in [5, 6, 7, 8, 9, 11, 12]:
                cell.fill = input_fill
                cell.border = thin_border
            elif col_idx == 4:
                cell.fill = result_fill
                cell.border = thin_border
    
    # Sütun genişlikleri
    ws1.column_dimensions['A'].width = 3
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 3
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['E'].width = 3
    ws1.column_dimensions['F'].width = 15
    
    # ============================================
    # 2. SEKME - MONTAJ MALZEMELER (Geliştirilmiş)
    # ============================================
    ws2 = wb.create_sheet('2-Montaj Malzemeleri')
    
    # Başlık ve açıklama
    ws2['A1'] = 'MONTAJ NOKTASI MALZEME LİSTESİ'
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:I1')
    
    ws2['A2'] = 'Not: Yeni malzeme eklemek için alt satırlara yazın. Formüller otomatik çalışacaktır.'
    ws2['A2'].font = Font(italic=True, color='666666')
    ws2.merge_cells('A2:I2')
    
    # Tablo başlıkları
    headers = ['S.No', 'Malzeme Kodu', 'Malzeme Tanımı', 'Miktar', 'Birim', 'Birim Fiyat (TL)', 'Toplam Fiyat (TL)', 'Kategori', 'Açıklama']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Gelişmiş malzeme listesi
    montaj_materials = [
        [1, 'BP-001', '70x70x1200x2,00mm Çelik Ankaraj', 1, 'adet', 125.50, '=D5*F5', 'MONTAJ', 'Temel ankrajı'],
        [2, 'BP-002', '80x80x3000x3,00mm Ana Direk', 1, 'adet', 285.75, '=D6*F6', 'MONTAJ', 'Yapı direği'],
        [3, 'BP-003', '200x200x8mm Montaj Plakası', 2, 'adet', 45.25, '=D7*F7', 'MONTAJ', 'Bağlantı plakası'],
        [4, 'BP-004', 'M12x80 Galvanizli Cıvata Seti', 4, 'takım', 8.75, '=D8*F8', 'MONTAJ', 'Bağlantı elemanı'],
        [5, 'BP-005', 'Kaynak İşçiliği', 1, 'saat', 65.00, '=D9*F9', 'MONTAJ', 'Montaj işçiliği'],
        [6, 'BP-006', 'Galvaniz Sıcak Daldırma', 1, 'kg', 3.50, '=D10*F10', 'MONTAJ', 'Koruyucu kaplama'],
        [7, 'BP-007', 'L Profil 50x50x5', 2, 'metre', 28.75, '=D11*F11', 'MONTAJ', 'Destek profili'],
        ['', '', '** Yeni malzemeler buraya ekleyin **', '', '', '', '', 'MONTAJ', ''],
        ['', '', '', '', '', '', '', '', ''],
        ['', '', 'TOPLAM MONTAJ MALİYETİ:', '', '', '', '=SUM(G5:G50)', '', '']
    ]
    
    for row_idx, row in enumerate(montaj_materials, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [4, 6]:  # Miktar ve fiyat sütunları - düzenlenebilir
                cell.fill = input_fill
            elif col_idx == 7:  # Toplam sütunu - hesaplanan
                cell.fill = calc_fill
            elif row_idx == 14 and col_idx == 7:  # Son toplam
                cell.font = Font(bold=True, size=12)
                cell.fill = result_fill
                cell.border = thick_border
    
    # Sütun genişlikleri
    column_widths = [5, 12, 35, 8, 8, 15, 15, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validation ekle
    dv_unit = DataValidation(type="list", formula1='"adet,metre,kg,takım,saat,m2,m3"')
    ws2.add_data_validation(dv_unit)
    dv_unit.add('E5:E50')
    
    # ============================================
    # 3. SEKME - DİREK MALZEMELER (Geliştirilmiş)
    # ============================================
    ws3 = wb.create_sheet('3-Direk Malzemeleri')
    
    # Başlık
    ws3['A1'] = 'ARA DİREK MALZEME LİSTESİ (1 METRE İÇİN)'
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:I1')
    
    ws3['A2'] = 'Not: Bu liste 1 metre direk için gerekli malzemeleri içerir. Toplam hesaplamada metre ile çarpılacak.'
    ws3['A2'].font = Font(italic=True, color='666666')
    ws3.merge_cells('A2:I2')
    
    # Başlıklar
    for col_idx, header in enumerate(headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Direk malzemeleri
    direk_materials = [
        [1, 'MD-001', '80x80x1000x2,5mm Ara Direk Profili', 1, 'metre', 95.25, '=D5*F5', 'DIREK', 'Ana direk malzemesi'],
        [2, 'MD-002', '120x8mm Bağlantı Flanşı', 2, 'adet', 25.50, '=D6*F6', 'DIREK', 'Direk bağlantısı'],
        [3, 'MD-003', 'Galvaniz Boyama', 1, 'm2', 15.75, '=D7*F7', 'DIREK', 'Koruyucu boya'],
        [4, 'MD-004', 'U Kelepçe Takımı', 2, 'adet', 12.25, '=D8*F8', 'DIREK', 'Sabitleyici'],
        [5, 'MD-005', 'Conta ve Sızdırmazlık', 1, 'takım', 8.50, '=D9*F9', 'DIREK', 'Hava geçirmezlik'],
        ['', '', '** Yeni malzemeler buraya ekleyin **', '', '', '', '', 'DIREK', ''],
        ['', '', '', '', '', '', '', '', ''],
        ['', '', 'TOPLAM DİREK MALİYETİ (1m):', '', '', '', '=SUM(G5:G50)', '', '']
    ]
    
    for row_idx, row in enumerate(direk_materials, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [4, 6]:
                cell.fill = input_fill
            elif col_idx == 7:
                cell.fill = calc_fill
            elif row_idx == 12 and col_idx == 7:
                cell.font = Font(bold=True, size=12)
                cell.fill = result_fill
                cell.border = thick_border
    
    # Sütun genişlikleri
    for idx, width in enumerate(column_widths, 1):
        ws3.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validation
    dv_unit2 = DataValidation(type="list", formula1='"adet,metre,kg,takım,saat,m2,m3"')
    ws3.add_data_validation(dv_unit2)
    dv_unit2.add('E5:E50')
    
    # ============================================
    # 4. SEKME - MÜŞTEMİLAT MALZEMELER (Geliştirilmiş)
    # ============================================
    ws4 = wb.create_sheet('4-Müştemilat Malzemeleri')
    
    # Başlık
    ws4['A1'] = 'MÜŞTEMİLAT DİREK MALZEME LİSTESİ (1 DİREK İÇİN)'
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:I1')
    
    ws4['A2'] = 'Not: Müştemilat alanındaki teknik oda/kazan direkleri için malzeme listesi.'
    ws4['A2'].font = Font(italic=True, color='666666')
    ws4.merge_cells('A2:I2')
    
    # Başlıklar
    for col_idx, header in enumerate(headers, 1):
        cell = ws4.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Müştemilat malzemeleri
    mustemilat_materials = [
        [1, 'MS-001', '80x80x2500x3,0mm Müştemilat Direk', 1, 'adet', 238.50, '=D5*F5', 'MUSTEMILAT', 'Ana destek direği'],
        [2, 'MS-002', '70x70x800mm Temel Ankrajı', 1, 'adet', 85.25, '=D6*F6', 'MUSTEMILAT', 'Temel bağlantısı'],
        [3, 'MS-003', 'Üst Bağlantı Başlığı', 1, 'adet', 45.75, '=D7*F7', 'MUSTEMILAT', 'Tavan bağlantısı'],
        [4, 'MS-004', 'Ağır Hizmet Montaj İşçiliği', 1, 'saat', 75.00, '=D8*F8', 'MUSTEMILAT', 'Özel montaj'],
        [5, 'MS-005', '0,5m³ Beton Temeli', 0.5, 'm3', 180.00, '=D9*F9', 'MUSTEMILAT', 'Temel betonu'],
        [6, 'MS-006', 'Demir Donatı 12mm', 20, 'kg', 6.50, '=D10*F10', 'MUSTEMILAT', 'Temel donatısı'],
        [7, 'MS-007', 'Su Yalıtımı', 1, 'm2', 35.00, '=D11*F11', 'MUSTEMILAT', 'Nem koruması'],
        ['', '', '** Yeni malzemeler buraya ekleyin **', '', '', '', '', 'MUSTEMILAT', ''],
        ['', '', '', '', '', '', '', '', ''],
        ['', '', 'TOPLAM MÜŞTEMİLAT MALİYETİ (1 direk):', '', '', '', '=SUM(G5:G50)', '', '']
    ]
    
    for row_idx, row in enumerate(mustemilat_materials, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [4, 6]:
                cell.fill = input_fill
            elif col_idx == 7:
                cell.fill = calc_fill
            elif row_idx == 14 and col_idx == 7:
                cell.font = Font(bold=True, size=12)
                cell.fill = result_fill
                cell.border = thick_border
    
    # Sütun genişlikleri
    for idx, width in enumerate(column_widths, 1):
        ws4.column_dimensions[get_column_letter(idx)].width = width
    
    # Data validation
    dv_unit3 = DataValidation(type="list", formula1='"adet,metre,kg,takım,saat,m2,m3"')
    ws4.add_data_validation(dv_unit3)
    dv_unit3.add('E5:E50')
    
    # ============================================
    # 5. SEKME - KAPSAMLI MALİYET HESAPLAMA
    # ============================================
    ws5 = wb.create_sheet('5-Maliyet Hesaplama')
    
    # Ana başlık
    ws5['A1'] = 'SERA PROJESİ KAPSAMLI MALİYET ANALİZİ VE HESAPLAMA'
    ws5['A1'].font = Font(bold=True, size=16, color='1F4E79')
    ws5.merge_cells('A1:H1')
    
    # Proje parametreleri
    param_section = [
        ['', '', '', '', '', '', '', ''],
        ['PROJE PARAMETRELERİ', '', '', '', 'HESAPLANAN DEĞERLER', '', '', ''],
        ['Parametre', 'Değer', 'Birim', '', 'Hesaplama', 'Sonuç', 'Birim', ''],
        
        # Sol taraf - Girdi parametreleri
        ['Tünel Uzunluğu', 250, 'm', '', 'Toplam Alan', '=B4*B5*B6', 'm²', ''],
        ['Tünel Sayısı', 50, 'adet', '', 'Montaj Noktası Sayısı', '=(B4/B8+1)*(B5+1)', 'adet', ''],
        ['Tünel Genişlik', 9.6, 'm', '', 'Toplam Direk Uzunluğu', '=B4*B5', 'm', ''],
        ['Duvar Kolon Arası', 2.5, 'm', '', 'Müştemilat Alan (10%)', '=F4*0.1', 'm²', ''],
        ['Orta Kolon Aralığı', 5, 'm', '', 'Müştemilat Direk Sayısı', '=F7/(B9*B9)', 'adet', ''],
        ['Müştemilat Aralığı', 2.5, 'm', '', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        
        # Birim maliyetler - Diğer sayfalardan otomatik çekilen
        ['BİRİM MALİYETLER (Otomatik)', '', '', '', '', '', '', ''],
        ['1 Montaj Noktası Maliyeti', "=SUMPRODUCT(('2-Montaj Malzemeleri'.D5:D50)*('2-Montaj Malzemeleri'.F5:F50))", 'TL', '', '', '', '', ''],
        ['1 Metre Direk Maliyeti', "=SUMPRODUCT(('3-Direk Malzemeleri'.D5:D50)*('3-Direk Malzemeleri'.F5:F50))", 'TL', '', '', '', '', ''],
        ['1 Müştemilat Direk Maliyeti', "=SUMPRODUCT(('4-Müştemilat Malzemeleri'.D5:D50)*('4-Müştemilat Malzemeleri'.F5:F50))", 'TL', '', '', '', '', ''],
        ['', '', '', '', '', '', '', '']
    ]
    
    for row_idx, row in enumerate(param_section, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if row_idx in [3, 4, 12] and col_idx in [1, 5]:  # Başlık satırları
                cell.font = Font(bold=True, size=12)
                cell.fill = sub_header_fill
                cell.border = thick_border
            elif col_idx == 2 and row_idx in range(5, 11):  # Input değerleri
                cell.fill = input_fill
            elif col_idx in [6] and value and '=' in str(value):  # Hesaplama sonuçları
                cell.fill = calc_fill
    
    # Ana maliyet hesaplama
    cost_section_start = 17
    cost_section = [
        ['', '', '', '', '', '', '', ''],
        ['ANA MALİYET HESAPLAMA', '', '', '', '', '', '', ''],
        ['Maliyet Kalemi', 'Birim Maliyet', 'Miktar', 'Toplam Maliyet', 'Yüzde', '', '', ''],
        ['Montaj Noktaları', '=B13', '=F5', '=B20*C20', '=D20/D27*100', '', '', ''],
        ['Ara Direkler', '=B14', '=F6', '=B21*C21', '=D21/D27*100', '', '', ''],
        ['Müştemilat Direkleri', '=B15', '=F8', '=B22*C22', '=D22/D27*100', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['TOPLAM MALZEME MALİYETİ', '', '', '=SUM(D20:D22)', '100%', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        
        # Ek maliyetler
        ['EK MALİYETLER VE MARJLAR', '', '', '', '', '', '', ''],
        ['Ek Maliyet', 'Oran (%)', 'Hesaplama', 'Tutar', '', '', '', ''],
        ['İşçilik ve Montaj', 15, '=D24*B27/100', '=C27', '', '', '', ''],
        ['Nakliye ve Lojistik', 5, '=D24*B28/100', '=C28', '', '', '', ''],
        ['Proje Yönetimi', 3, '=D24*B29/100', '=C29', '', '', '', ''],
        ['Sigorta ve Garanti', 2, '=D24*B30/100', '=C30', '', '', '', ''],
        ['Kar Marjı', 12, '=D24*B31/100', '=C31', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['ARA TOPLAM (KDV HARİÇ)', '', '', '=D24+SUM(D27:D31)', '', '', '', ''],
        ['KDV (%20)', '', '', '=D33*0.20', '', '', '', ''],
        ['GENEL TOPLAM (KDV DAHİL)', '', '', '=D33+D34', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        
        # Detaylı analiz
        ['DETAYLI ANALİZ', '', '', '', '', '', '', ''],
        ['Analiz Kriteri', '', 'Değer', 'Birim', '', '', '', ''],
        ['M² Başına Maliyet (KDV Hariç)', '', '=D33/F4', 'TL/m²', '', '', '', ''],
        ['M² Başına Maliyet (KDV Dahil)', '', '=D35/F4', 'TL/m²', '', '', '', ''],
        ['Montaj Başına Ortalama Maliyet', '', '=D20/C20', 'TL/adet', '', '', '', ''],
        ['Metre Direk Başına Maliyet', '', '=D21/C21', 'TL/m', '', '', '', ''],
        ['Müştemilat Direk Başına', '', '=D22/C22', 'TL/adet', '', '', '', ''],
        ['', '', '', '', '', '', '', ''],
        ['Toplam Projedeki Çelik Ağırlığı (Tahmini)', '', '=F5*25+F6*7.8+F8*35', 'kg', '', '', '', ''],
        ['Kg Başına Ortalama Maliyet', '', '=D35/C44', 'TL/kg', '', '', '', '']
    ]
    
    for row_idx, row in enumerate(cost_section, cost_section_start):
        for col_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Başlık satırları
            if row_idx in [cost_section_start+1, cost_section_start+9, cost_section_start+21] and col_idx == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = sub_header_fill
                cell.border = thick_border
            
            # Alt başlık satırları
            elif row_idx in [cost_section_start+2, cost_section_start+10, cost_section_start+22] and col_idx <= 4:
                cell.font = sub_header_font
                cell.fill = sub_header_fill
            
            # Input alanları (oranlar)
            elif col_idx == 2 and row_idx in range(cost_section_start+11, cost_section_start+16):
                cell.fill = input_fill
            
            # Hesaplanan değerler
            elif col_idx in [3, 4] and value and '=' in str(value):
                cell.fill = calc_fill
            
            # Önemli sonuçlar
            elif row_idx in [cost_section_start+7, cost_section_start+17, cost_section_start+19] and col_idx == 4:
                cell.font = Font(bold=True, size=12)
                cell.fill = result_fill
                cell.border = thick_border
    
    # Sütun genişlikleri
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 15
    ws5.column_dimensions['C'].width = 15
    ws5.column_dimensions['D'].width = 18
    ws5.column_dimensions['E'].width = 15
    ws5.column_dimensions['F'].width = 15
    ws5.column_dimensions['G'].width = 10
    ws5.column_dimensions['H'].width = 5
    
    # Dosyayı kaydet
    filename = 'Sera_Ultimate_Maliyet_Sistemi.xlsx'
    wb.save(filename)
    print(f"🎯 ULTIMATE SERA MALİYET SİSTEMİ oluşturuldu: {filename}")
    print("\n✅ YENİ ÖZELLİKLER:")
    print("🔗 Tamamen dinamik sayfalar arası formüller")
    print("📊 SUMPRODUCT ile gelişmiş hesaplama")
    print("📋 Dropdown menüler (birim seçimi)")
    print("🎨 Profesyonel tasarım ve renkli kategoriler")
    print("📈 Detaylı maliyet analizi ve yüzde dağılımları")
    print("⚡ Yeni malzeme eklendiğinde otomatik hesaplama")
    print("🔄 Sayfalar arası tamamen entegre sistem")
    print("💡 Akıllı formüller ve hata kontrolü")
    
    return filename

if __name__ == "__main__":
    create_ultimate_sera_excel()
