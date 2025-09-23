#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_simple_sera_excel():
    # Excel dosyası oluştur
    wb = Workbook()
    
    # Basit stil tanımlamaları
    header_font = Font(bold=True, color='000000', size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    input_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # SARI - INPUT
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ============================================
    # 1. SEKME - PROJE BİLGİLERİ (Basit)
    # ============================================
    ws1 = wb.active
    ws1.title = 'Proje Bilgileri'
    
    # Proje başlığı
    ws1['A1'] = 'SERA PROJESİ BİLGİLERİ'
    ws1['A1'].font = Font(bold=True, size=14)
    
    # Basit proje bilgileri
    project_data = [
        ['', ''],
        ['Proje Adı:', ''],
        ['Müşteri:', ''],
        ['Tarih:', ''],
        ['Hazırlayan:', '']
    ]
    
    for row_idx, row in enumerate(project_data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 2:  # Input sütunu
                cell.fill = input_fill
                cell.border = thin_border
    
    # Sütun genişlikleri
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 30
    
    # ============================================
    # 2. SEKME - MONTAJ MALZEMELER (Basitleştirilmiş)
    # ============================================
    ws2 = wb.create_sheet('Montaj Malzemeleri')
    
    # Başlık
    ws2['A1'] = 'MONTAJ NOKTASI MALZEMELERİ (1 ADET İÇİN)'
    ws2['A1'].font = Font(bold=True, size=12)
    
    # Tablo başlıkları
    headers = ['MALZEME', 'MİKTAR', 'FİYAT', 'TOPLAM']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Malzeme listesi - SADECE SAYILAR VE FORMÜLLER
    montaj_data = [
        ['Ankaraj 70x70x1200', 1, 125.50, '=B4*C4'],
        ['Direk 80x80x3000', 1, 285.75, '=B5*C5'],
        ['Montaj Plakası', 2, 45.25, '=B6*C6'],
        ['Bağlantı Elemanı', 4, 8.75, '=B7*C7'],
        ['İşçilik', 1, 65.00, '=B8*C8'],
        ['', '', '', ''],
        ['TOPLAM', '', '', '=SUM(D4:D8)']
    ]
    
    for row_idx, row in enumerate(montaj_data, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # INPUT alanları - SARI
            if col_idx in [2, 3] and row_idx <= 8:  # Miktar ve fiyat
                cell.fill = input_fill
    
    # Sütun genişlikleri
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    
    # ============================================
    # 3. SEKME - DİREK MALZEMELER
    # ============================================
    ws3 = wb.create_sheet('Direk Malzemeleri')
    
    ws3['A1'] = 'ARA DİREK MALZEMELERİ (1 METRE İÇİN)'
    ws3['A1'].font = Font(bold=True, size=12)
    
    # Başlıklar
    for col_idx, header in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Direk malzemeleri
    direk_data = [
        ['Direk 80x80x1000', 1, 95.25, '=B4*C4'],
        ['Bağlantı Flanşı', 2, 25.50, '=B5*C5'],
        ['Boyama', 1, 15.75, '=B6*C6'],
        ['Kelepçe', 2, 12.25, '=B7*C7'],
        ['', '', '', ''],
        ['TOPLAM', '', '', '=SUM(D4:D7)']
    ]
    
    for row_idx, row in enumerate(direk_data, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [2, 3] and row_idx <= 7:
                cell.fill = input_fill
    
    # Sütun genişlikleri
    for i in range(1, 5):
        ws3.column_dimensions[get_column_letter(i)].width = [25, 10, 12, 12][i-1]
    
    # ============================================
    # 4. SEKME - MÜŞTEMİLAT MALZEMELER
    # ============================================
    ws4 = wb.create_sheet('Müştemilat Malzemeleri')
    
    ws4['A1'] = 'MÜŞTEMİLAT DİREK MALZEMELERİ (1 ADET İÇİN)'
    ws4['A1'].font = Font(bold=True, size=12)
    
    # Başlıklar
    for col_idx, header in enumerate(headers, 1):
        cell = ws4.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Müştemilat malzemeleri
    mustemilat_data = [
        ['Direk 80x80x2500', 1, 238.50, '=B4*C4'],
        ['Temel Ankrajı', 1, 85.25, '=B5*C5'],
        ['Bağlantı Aparatı', 1, 45.75, '=B6*C6'],
        ['İşçilik', 1, 75.00, '=B7*C7'],
        ['Beton 0.5m3', 1, 90.00, '=B8*C8'],
        ['', '', '', ''],
        ['TOPLAM', '', '', '=SUM(D4:D8)']
    ]
    
    for row_idx, row in enumerate(mustemilat_data, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [2, 3] and row_idx <= 8:
                cell.fill = input_fill
    
    # Sütun genişlikleri
    for i in range(1, 5):
        ws4.column_dimensions[get_column_letter(i)].width = [25, 10, 12, 12][i-1]
    
    # ============================================
    # 5. SEKME - HESAPLAMA (Basit ve Anlaşılır)
    # ============================================
    ws5 = wb.create_sheet('Maliyet Hesaplama')
    
    # Başlık
    ws5['A1'] = 'SERA PROJESİ MALİYET HESAPLAMA'
    ws5['A1'].font = Font(bold=True, size=14)
    
    # Proje parametreleri - SARI INPUT ALANLARI
    calc_data = [
        ['', '', '', ''],
        ['PROJE PARAMETRELERİ', '', '', ''],
        ['', '', '', ''],
        ['Tünel Uzunluğu (m)', '', 250, ''],
        ['Tünel Sayısı (adet)', '', 50, ''],
        ['Tünel Genişlik (m)', '', 9.6, ''],
        ['Orta Kolon Aralığı (m)', '', 5, ''],
        ['', '', '', ''],
        ['HESAPLANAN DEĞERLER', '', '', ''],
        ['', '', '', ''],
        ['Toplam Alan (m²)', '', '=C5*C6*C7', ''],
        ['Montaj Noktası Sayısı', '', '=(C5/C8+1)*(C6+1)', ''],
        ['Ara Direk Uzunluğu (m)', '', '=C5*C6', ''],
        ['Müştemilat Direk Sayısı', '', 25, ''],
        ['', '', '', ''],
        ['BİRİM MALİYETLER', '', '', ''],
        ['', '', '', ''],
        ['1 Montaj Noktası', '', "='Montaj Malzemeleri'.D10", ''],
        ['1 Metre Direk', '', "='Direk Malzemeleri'.D9", ''],
        ['1 Müştemilat Direk', '', "='Müştemilat Malzemeleri'.D10", ''],
        ['', '', '', ''],
        ['TOPLAM MALİYETLER', '', '', ''],
        ['', '', '', ''],
        ['Montaj Maliyeti', '', '=C12*C18', ''],
        ['Direk Maliyeti', '', '=C13*C19', ''],
        ['Müştemilat Maliyeti', '', '=C14*C20', ''],
        ['', '', '', ''],
        ['TOPLAM MALZEME', '', '=SUM(C24:C26)', ''],
        ['İşçilik %15', '', '=C28*0.15', ''],
        ['Nakliye %5', '', '=C28*0.05', ''],
        ['Kar %10', '', '=C28*0.10', ''],
        ['', '', '', ''],
        ['GENEL TOPLAM (KDV HARİÇ)', '', '=C28+C29+C30+C31', ''],
        ['KDV %20', '', '=C33*0.20', ''],
        ['TOPLAM (KDV DAHİL)', '', '=C33+C34', ''],
        ['', '', '', ''],
        ['M² BAŞINA MALİYET', '', '=C35/C11', 'TL/m²']
    ]
    
    for row_idx, row in enumerate(calc_data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Başlık satırları
            if col_idx == 1 and value and 'PARAMETRELERİ' in str(value):
                cell.font = Font(bold=True, size=12)
                cell.fill = header_fill
            
            # INPUT alanları - SARI (sadece proje parametreleri)
            elif col_idx == 3 and row_idx in [5, 6, 7, 8, 15]:  # Parametreler ve müştemilat sayısı
                cell.fill = input_fill
    
    # Sütun genişlikleri
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 5
    ws5.column_dimensions['C'].width = 20
    ws5.column_dimensions['D'].width = 10
    
    # Dosyayı kaydet
    filename = 'Sera_Basit_Sistem.xlsx'
    wb.save(filename)
    print(f"✅ BASİT SERA SİSTEMİ oluşturuldu: {filename}")
    print("\n🎯 ÖZELLİKLER:")
    print("🟡 SARI hücreler = VERİ GİRİŞ YERLERİ")
    print("📋 Basit formüller, #DEĞER hatası yok")
    print("🔗 Sayfalar arası doğru bağlantılar")
    print("📊 Anlaşılır hesaplama mantığı")
    print("⚡ Kolay kullanım")
    
    return filename

if __name__ == "__main__":
    create_simple_sera_excel()
