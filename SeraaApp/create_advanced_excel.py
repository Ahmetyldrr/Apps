#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_advanced_sera_excel():
    # Excel dosyası oluştur
    wb = Workbook()
    
    # Stil tanımlamaları
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    sub_header_font = Font(bold=True, color='000000', size=11)
    sub_header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ============================================
    # 1. SEKME - PROJE BİLGİLERİ
    # ============================================
    ws1 = wb.active
    ws1.title = '1-Proje Bilgileri'
    
    # Proje bilgileri başlık
    ws1['A1'] = 'SERA PROJESİ GENEL BİLGİLER'
    ws1['A1'].font = Font(bold=True, size=16, color='2F5597')
    ws1.merge_cells('A1:E1')
    
    # Proje detayları
    proje_info = [
        ['', '', '', '', ''],
        ['Proje Adı:', '', '', '', ''],
        ['Proje Kodu:', '', '', '', ''],
        ['Müşteri:', '', '', '', ''],
        ['Tarih:', '', '', '', ''],
        ['Hazırlayan:', '', '', '', ''],
        ['Onaylayan:', '', '', '', ''],
        ['', '', '', '', ''],
        ['PROJE DURUMU', '', '', '', ''],
        ['Durum:', 'Hazırlanıyor', '', '', ''],
        ['Revizyon:', '1.0', '', '', ''],
        ['Son Güncelleme:', '=TODAY()', '', '', '']
    ]
    
    for row_idx, row in enumerate(proje_info, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx+1, column=col_idx, value=value)
            if col_idx == 1 and value and ':' in str(value):
                cell.font = sub_header_font
            elif col_idx == 2 and row_idx in [1, 2, 3, 4, 5, 6]:
                cell.fill = input_fill
                cell.border = thin_border
    
    # Sütun genişlikleri
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    
    # ============================================
    # 2. SEKME - MONTAJ NOKTASI MALZEMELER
    # ============================================
    ws2 = wb.create_sheet('2-Montaj Malzemeleri')
    
    # Başlık
    headers = ['SIRA', 'MALZEME KODU', 'MALZEME TANIMI', 'MİKTAR', 'BİRİM', 'BİRİM FİYAT', 'TOPLAM FİYAT', 'KATEGORİ']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Örnek malzemeler - MONTAJ kategorisinde
    montaj_materials = [
        [1, 'BP-01', '70*70*1200*2,00 mm Ankaraj', 1, 'adet', 125.50, '=D2*F2', 'MONTAJ'],
        [2, 'BP-02', '80*80*3000*3,00 mm Direk', 1, 'adet', 285.75, '=D3*F3', 'MONTAJ'],
        [3, 'BP-03', 'Montaj Plakası 200*200*8mm', 2, 'adet', 45.25, '=D4*F4', 'MONTAJ'],
        [4, 'BP-04', 'Bağlantı Elemanı M12*80', 4, 'adet', 8.75, '=D5*F5', 'MONTAJ'],
        [5, 'BP-05', 'Kaynak İşçiliği', 1, 'saat', 65.00, '=D6*F6', 'MONTAJ'],
        [6, 'BP-06', 'Galvaniz Boyama', 1, 'adet', 35.00, '=D7*F7', 'MONTAJ'],
        ['', '', '', '', '', '', '', ''],
        ['', 'Yeni malzemeler buraya eklenebilir...', '', '', '', '', '', '']
    ]
    
    for row_idx, row in enumerate(montaj_materials, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in [4, 6]:  # Miktar ve fiyat sütunları
                cell.fill = input_fill
            elif col_idx == 7:  # Toplam sütunu
                cell.fill = calc_fill
    
    # Sütun genişlikleri
    column_widths = [8, 15, 35, 8, 8, 12, 12, 12]
    for idx, width in enumerate(column_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width
    
    # ============================================
    # 3. SEKME - ARA DİREK MALZEMELER
    # ============================================
    ws3 = wb.create_sheet('3-Direk Malzemeleri')
    
    # Başlık
    for col_idx, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Örnek malzemeler - DIREK kategorisinde
    direk_materials = [
        [1, 'MD-01', '80*80*1000*2,5mm Ara Direk', 1, 'metre', 95.25, '=D2*F2', 'DIREK'],
        [2, 'MD-02', 'Bağlantı Flanşı', 2, 'adet', 25.50, '=D3*F3', 'DIREK'],
        [3, 'MD-03', 'Galvaniz Boyama', 1, 'm2', 15.75, '=D4*F4', 'DIREK'],
        [4, 'MD-04', 'Montaj Kelepçesi', 2, 'adet', 12.25, '=D5*F5', 'DIREK'],
        ['', '', '', '', '', '', '', ''],
        ['', 'Yeni malzemeler buraya eklenebilir...', '', '', '', '', '', '']
    ]
    
    for row_idx, row in enumerate(direk_materials, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in [4, 6]:  # Miktar ve fiyat sütunları
                cell.fill = input_fill
            elif col_idx == 7:  # Toplam sütunu
                cell.fill = calc_fill
    
    # Sütun genişlikleri
    for idx, width in enumerate(column_widths, 1):
        ws3.column_dimensions[get_column_letter(idx)].width = width
    
    # ============================================
    # 4. SEKME - MÜŞTEMİLAT MALZEMELER
    # ============================================
    ws4 = wb.create_sheet('4-Müştemilat Malzemeleri')
    
    # Başlık
    for col_idx, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Örnek malzemeler - MUSTEMILAT kategorisinde
    mustemilat_materials = [
        [1, 'MS-01', '80*80*2500*3,0mm Müştemilat Direk', 1, 'adet', 238.50, '=D2*F2', 'MUSTEMILAT'],
        [2, 'MS-02', 'Temel Ankrajı 70*70*800mm', 1, 'adet', 85.25, '=D3*F3', 'MUSTEMILAT'],
        [3, 'MS-03', 'Üst Bağlantı Aparatı', 1, 'adet', 45.75, '=D4*F4', 'MUSTEMILAT'],
        [4, 'MS-04', 'Montaj İşçiliği', 1, 'saat', 75.00, '=D5*F5', 'MUSTEMILAT'],
        [5, 'MS-05', 'Beton Temeli', 0.5, 'm3', 180.00, '=D6*F6', 'MUSTEMILAT'],
        ['', '', '', '', '', '', '', ''],
        ['', 'Yeni malzemeler buraya eklenebilir...', '', '', '', '', '', '']
    ]
    
    for row_idx, row in enumerate(mustemilat_materials, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in [4, 6]:  # Miktar ve fiyat sütunları
                cell.fill = input_fill
            elif col_idx == 7:  # Toplam sütunu
                cell.fill = calc_fill
    
    # Sütun genişlikleri
    for idx, width in enumerate(column_widths, 1):
        ws4.column_dimensions[get_column_letter(idx)].width = width
    
    # ============================================
    # 5. SEKME - HESAPLAMA VE MALİYETLENDİRME
    # ============================================
    ws5 = wb.create_sheet('5-Maliyet Hesaplama')
    
    # Ana başlık
    ws5['A1'] = 'SERA PROJESİ MALİYET HESAPLAMA VE ANALİZ'
    ws5['A1'].font = Font(bold=True, size=16, color='2F5597')
    ws5.merge_cells('A1:F1')
    
    # Proje parametreleri bölümü
    param_data = [
        ['', '', '', '', '', ''],
        ['PROJE PARAMETRELERİ', '', '', '', '', ''],
        ['Parametre', 'Değer', 'Birim', '', 'Hesaplama', 'Sonuç'],
        ['Tünel Uzunluğu', 250, 'm', '', '', ''],
        ['Tünel Sayısı', 50, 'adet', '', '', ''],
        ['Tünel Genişlik', 9.6, 'm', '', '', ''],
        ['Duvar Kolon Arası', 2.5, 'm', '', '', ''],
        ['Orta Kolon Aralığı', 5, 'm', '', '', ''],
        ['Müştemilat Direk Aralığı', 2.5, 'm', '', '', ''],
        ['', '', '', '', '', ''],
        ['HESAPLANAN DEĞERLER', '', '', '', '', ''],
        ['Toplam Alan', '', 'm²', '', '=B5*B6*B7', '=E12'],
        ['Montaj Noktası Sayısı', '', 'adet', '', '=(B5/B9+1)*(B6+1)', '=E13'],
        ['Toplam Direk Uzunluğu', '', 'm', '', '=B5*B6', '=E14'],
        ['Müştemilat Alan', '', 'm²', '', '=F12*0.1', '=E15'],
        ['Müştemilat Direk Sayısı', '', 'adet', '', '=E15/B10/B10', '=E16']
    ]
    
    for row_idx, row in enumerate(param_data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if row_idx in [3, 4, 11] and col_idx <= 3:  # Başlık satırları
                cell.font = sub_header_font
                cell.fill = sub_header_fill
            elif col_idx == 2 and row_idx in range(5, 11):  # Input değerleri
                cell.fill = input_fill
            elif col_idx in [5, 6] and row_idx >= 12:  # Hesaplama formülleri
                cell.fill = calc_fill
    
    # Maliyet hesaplama bölümü
    cost_start_row = 19
    cost_data = [
        ['', '', '', '', '', ''],
        ['MALİYET HESAPLAMA', '', '', '', '', ''],
        ['Kategori', 'Birim Maliyet', 'Miktar', 'Toplam Maliyet', 'Açıklama', ''],
        
        # Dinamik SUMIF formülleri ile diğer sayfalardan veri çekme
        ['1 Montaj Noktası', 
         '=SUMIF(\'2-Montaj Malzemeleri\'.H:H,"MONTAJ",\'2-Montaj Malzemeleri\'.G:G)', 
         '=F13', 
         '=B23*C23', 
         'Otomatik hesaplanan', ''],
        
        ['1 Metre Direk', 
         '=SUMIF(\'3-Direk Malzemeleri\'.H:H,"DIREK",\'3-Direk Malzemeleri\'.G:G)', 
         '=F14', 
         '=B24*C24', 
         'Otomatik hesaplanan', ''],
        
        ['1 Müştemilat Direk', 
         '=SUMIF(\'4-Müştemilat Malzemeleri\'.H:H,"MUSTEMILAT",\'4-Müştemilat Malzemeleri\'.G:G)', 
         '=F16', 
         '=B25*C25', 
         'Otomatik hesaplanan', ''],
        
        ['', '', '', '', '', ''],
        ['TOPLAM MALZEME MALİYETİ', '', '', '=SUM(D23:D25)', '', ''],
        ['', '', '', '', '', ''],
        
        # Ek maliyetler
        ['Ek Maliyetler', 'Oran %', 'Tutar', 'Toplam', '', ''],
        ['İşçilik', 15, '=D28*B31/100', '=C31', '', ''],
        ['Nakliye', 5, '=D28*B32/100', '=C32', '', ''],
        ['Montaj', 8, '=D28*B33/100', '=C33', '', ''],
        ['Kar Marjı', 12, '=D28*B34/100', '=C34', '', ''],
        ['', '', '', '', '', ''],
        ['TOPLAM (KDV HARİÇ)', '', '', '=D28+SUM(D31:D34)', '', ''],
        ['KDV (%20)', '', '', '=D36*0.20', '', ''],
        ['GENEL TOPLAM', '', '', '=D36+D37', '', ''],
        ['', '', '', '', '', ''],
        
        # Analiz
        ['ANALİZ', '', '', '', '', ''],
        ['M² Başına Maliyet (KDV Hariç)', '', '', '=D36/F12', 'TL/m²', ''],
        ['M² Başına Maliyet (KDV Dahil)', '', '', '=D38/F12', 'TL/m²', ''],
        ['Montaj Başına Ortalama', '', '', '=D23', 'TL/adet', ''],
        ['Direk Metre Başına', '', '', '=B24', 'TL/m', '']
    ]
    
    for row_idx, row in enumerate(cost_data, cost_start_row):
        for col_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if row_idx in [cost_start_row+1, cost_start_row+2, cost_start_row+9, cost_start_row+19] and col_idx <= 4:
                cell.font = sub_header_font
                cell.fill = sub_header_fill
            elif col_idx == 2 and row_idx in [cost_start_row+10, cost_start_row+11, cost_start_row+12, cost_start_row+13]:
                cell.fill = input_fill
            elif col_idx == 4 and 'SUM' in str(value):
                cell.font = Font(bold=True)
                cell.fill = calc_fill
    
    # Sütun genişlikleri
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 15
    ws5.column_dimensions['C'].width = 12
    ws5.column_dimensions['D'].width = 15
    ws5.column_dimensions['E'].width = 20
    ws5.column_dimensions['F'].width = 12
    
    # Dosyayı kaydet
    filename = 'Sera_Gelismis_Maliyet_Sistemi.xlsx'
    wb.save(filename)
    print(f"Gelişmiş Excel dosyası başarıyla oluşturuldu: {filename}")
    print("Özellikler:")
    print("✓ Sayfalar arası dinamik formüller")
    print("✓ SUMIF ile kategori bazlı otomatik toplama")
    print("✓ Yeni malzeme eklendiğinde otomatik hesaplama")
    print("✓ Gelişmiş maliyet analizi")
    print("✓ Dinamik proje parametreleri")
    
    return filename

if __name__ == "__main__":
    create_advanced_sera_excel()
