#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_ultimate_professional_excel():
    # Excel dosyası oluştur
    wb = Workbook()
    
    # Profesyonel renkler
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')  
    input_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')   # SARI INPUT
    calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')    
    result_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  
    
    # Fontlar
    title_font = Font(bold=True, size=16, color='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), 
                         top=Side(style='medium'), bottom=Side(style='medium'))
    
    # ============================================
    # 1. SEKME - PROJE YÖNETİMİ
    # ============================================
    ws1 = wb.active
    ws1.title = 'Dashboard'
    
    ws1['A1'] = 'SERA PROJESİ KONTROL PANELİ'
    ws1['A1'].font = title_font
    
    # Basit proje bilgileri
    dashboard_data = [
        ['', '', ''],
        ['PROJE BİLGİLERİ', '', ''],
        ['Proje Adı:', '', ''],
        ['Müşteri:', '', ''],
        ['Tarih:', '=TODAY()', ''],
        ['', '', ''],
        ['HIZLI DURUM', '', ''],
        ['Toplam Maliyet:', "=Hesaplama!C37", '₺'],
        ['M² Başına:', "=Hesaplama!C37/SUM(Hesaplama!C11:C13)", '₺/adet'],
        ['Durum:', 'Hesaplanıyor', '']
    ]
    
    for row_idx, row in enumerate(dashboard_data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if 'BİLGİLERİ' in str(value) or 'DURUM' in str(value):
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.font = Font(bold=True, color='FFFFFF')
            elif col_idx == 3 and row_idx in [4, 5]:
                cell.fill = input_fill
            elif '=' in str(value):
                cell.fill = calc_fill
    
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 5
    ws1.column_dimensions['C'].width = 25
    
    # ============================================
    # 2. SEKME - MONTAJ MALZEMELER
    # ============================================
    ws2 = wb.create_sheet('Montaj')
    
    ws2['A1'] = 'MONTAJ NOKTASI MALZEMELERİ (1 ADET İÇİN)'
    ws2['A1'].font = title_font
    
    # Başlıklar
    headers = ['Malzeme', 'Miktar', 'Fiyat (₺)', 'Toplam (₺)']
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
    
    # Malzemeler
    montaj_data = [
        ['Çelik Ankaraj 70x70x1200', 1, 145.50, '=B4*C4'],
        ['Ana Direk 80x80x3000', 1, 325.75, '=B5*C5'],
        ['Montaj Plakası 200x200', 2, 65.25, '=B6*C6'],
        ['Cıvata Seti M12x80', 1, 28.75, '=B7*C7'],
        ['Kaynak İşçiliği', 2, 75.00, '=B8*C8'],
        ['Galvaniz Kaplama', 1, 55.00, '=B9*C9'],
        ['Destek Profili L50x50', 1.5, 42.25, '=B10*C10'],
        ['', '', '', ''],
        ['TOPLAM (1 MONTAJ)', '', '', '=SUM(D4:D10)']
    ]
    
    for row_idx, row in enumerate(montaj_data, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [2, 3] and row_idx <= 10:  # INPUT alanları
                cell.fill = input_fill
            elif col_idx == 4 and '=' in str(value):
                cell.fill = calc_fill
            elif row_idx == 12:
                cell.font = Font(bold=True)
                cell.fill = result_fill
    
    for i in range(1, 5):
        ws2.column_dimensions[get_column_letter(i)].width = [30, 10, 12, 12][i-1]
    
    # ============================================
    # 3. SEKME - DİREK MALZEMELER
    # ============================================
    ws3 = wb.create_sheet('Direk')
    
    ws3['A1'] = 'ARA DİREK MALZEMELERİ (1 METRE İÇİN)'
    ws3['A1'].font = title_font
    
    # Başlıklar
    for col_idx, header in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
    
    # Direk malzemeleri
    direk_data = [
        ['Ara Direk 80x80x1000', 1, 115.25, '=B4*C4'],
        ['Flanş Bağlantısı', 2, 35.50, '=B5*C5'],
        ['Toz Boya', 0.8, 25.75, '=B6*C6'],
        ['Montaj Kelepçesi', 2, 18.25, '=B7*C7'],
        ['Conta Sistemi', 1, 12.50, '=B8*C8'],
        ['', '', '', ''],
        ['TOPLAM (1 METRE)', '', '', '=SUM(D4:D8)']
    ]
    
    for row_idx, row in enumerate(direk_data, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [2, 3] and row_idx <= 8:
                cell.fill = input_fill
            elif col_idx == 4 and '=' in str(value):
                cell.fill = calc_fill
            elif row_idx == 10:
                cell.font = Font(bold=True)
                cell.fill = result_fill
    
    for i in range(1, 5):
        ws3.column_dimensions[get_column_letter(i)].width = [30, 10, 12, 12][i-1]
    
    # ============================================
    # 4. SEKME - MÜŞTEMİLAT MALZEMELER
    # ============================================
    ws4 = wb.create_sheet('Müştemilat')
    
    ws4['A1'] = 'MÜŞTEMİLAT DİREK MALZEMELERİ (1 ADET İÇİN)'
    ws4['A1'].font = title_font
    
    # Başlıklar
    for col_idx, header in enumerate(headers, 1):
        cell = ws4.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
    
    # Müştemilat malzemeleri
    mustemilat_data = [
        ['Ağır Direk 80x80x2500', 1, 285.50, '=B4*C4'],
        ['Temel Ankrajı 70x70x800', 1, 125.25, '=B5*C5'],
        ['Taşıma Başlığı', 1, 85.75, '=B6*C6'],
        ['Özel Montaj İşçiliği', 3, 95.00, '=B7*C7'],
        ['Beton Temeli 0.6m³', 0.6, 220.00, '=B8*C8'],
        ['Donatı Φ14', 25, 12.50, '=B9*C9'],
        ['Su Yalıtımı', 1.2, 45.00, '=B10*C10'],
        ['', '', '', ''],
        ['TOPLAM (1 MÜŞTEMİLAT)', '', '', '=SUM(D4:D10)']
    ]
    
    for row_idx, row in enumerate(mustemilat_data, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx in [2, 3] and row_idx <= 10:
                cell.fill = input_fill
            elif col_idx == 4 and '=' in str(value):
                cell.fill = calc_fill
            elif row_idx == 12:
                cell.font = Font(bold=True)
                cell.fill = result_fill
    
    for i in range(1, 5):
        ws4.column_dimensions[get_column_letter(i)].width = [30, 10, 12, 12][i-1]
    
    # ============================================
    # 5. SEKME - ANA HESAPLAMA MERKEZİ
    # ============================================
    ws5 = wb.create_sheet('Hesaplama')
    
    ws5['A1'] = 'SERA PROJESİ MALİYET HESAPLAMA MERKEZİ'
    ws5['A1'].font = title_font
    
    # Ana hesaplama tablosu - Gelişmiş formüllerle
    calc_data = [
        ['', '', '', ''],
        ['BİRİM MALİYET HESAPLAMA', '', '', ''],
        ['', '', '', ''],
        ['1 Montaj Noktası Maliyeti', '', '=Montaj!D12', '₺'],
        ['1 Metre Direk Maliyeti', '', '=Direk!D10', '₺'],  
        ['1 Müştemilat Direk Maliyeti', '', '=Müştemilat!D12', '₺'],
        ['', '', '', ''],
        ['PROJE GİRDİ PARAMETRELERİ', '', '', ''],
        ['', '', '', ''],
        ['Montaj Noktası Adedi', '', 10, 'Kaç adet montaj'],
        ['Direk Uzunluğu (metre)', '', 25, 'Kaç metre direk'],
        ['Müştemilat Adedi', '', 3, 'Kaç adet müştemilat'],
        ['', '', '', ''],
        ['BASİT HESAPLAMA', '', '', ''],
        ['', '', '', ''],
        ['Montaj Maliyeti', '', '=C5*C11', '₺'],
        ['Direk Maliyeti', '', '=C6*C12', '₺'],
        ['Müştemilat Maliyeti', '', '=C7*C13', '₺'],
        ['', '', '', ''],
        ['TOPLAM MALZEME MALİYETİ', '', '=SUM(C17:C19)', '₺'],
        ['', '', '', ''],
        ['EK MALİYET HESAPLAMALARI', '', '', ''],
        ['', '', '', ''],
        ['İşçilik Oranı (%)', '', '=IF(C21<50000,18,IF(C21<200000,15,12))', 'Dinamik oran'],
        ['İşçilik Sabit Bedel (₺)', '', '=IF(C21<50000,5000,IF(C21<200000,8000,12000))', 'Sabit ekleme'],
        ['Nakliye Oranı (%)', '', '=MIN(8,MAX(3,C21/50000))', 'Dinamik oran'],
        ['Nakliye Sabit Bedel (₺)', '', '=IF(C21<30000,2500,IF(C21<100000,4000,6500))', 'Sabit ekleme'],
        ['Kar Marjı (%)', '', '=IF(C21<100000,15,IF(C21<500000,12,10))', 'Dinamik oran'],
        ['Kar Sabit Bedel (₺)', '', '=IF(C21<75000,3000,IF(C21<300000,5500,8000))', 'Sabit ekleme'],
        ['', '', '', ''],
        ['İşçilik Maliyeti', '', '=ROUND(C21*C25/100+C26,2)', '₺'],
        ['Nakliye Maliyeti', '', '=ROUND(C21*C27/100+C28,2)', '₺'],
        ['Kar Marjı', '', '=ROUND(C21*C29/100+C30,2)', '₺'],
        ['', '', '', ''],
        ['ARA TOPLAM (KDV HARİÇ)', '', '=C21+C32+C33+C34', '₺'],
        ['KDV (%20)', '', '=ROUND(C35*0.20,2)', '₺'],
        ['GENEL TOPLAM (KDV DAHİL)', '', '=C35+C36', '₺'],
        ['', '', '', ''],
        ['BİRİM FİYAT ANALİZİ', '', '', ''],
        ['', '', '', ''],
        ['M² Başına (KDV Hariç)', '', '=ROUND(C35/SUM(C11:C13),2)', '₺/adet'],
        ['M² Başına (KDV Dahil)', '', '=ROUND(C37/SUM(C11:C13),2)', '₺/adet'],
        ['Karlılık Oranı', '', '=ROUND(C34/C21*100,1)', '%']
    ]
    
    for row_idx, row in enumerate(calc_data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Başlık satırları
            if col_idx == 1 and ('PARAMETRELERİ' in str(value) or 'BOYUTLAR' in str(value) or 'MALİYETLER' in str(value)):
                cell.font = Font(bold=True, size=12)
                cell.fill = header_fill
                cell.font = Font(bold=True, color='FFFFFF')
            
            # INPUT alanları - SARI  
            elif col_idx == 3 and row_idx in [11, 12, 13]:  # 3 adet input: montaj, direk, müştemilat (doğru satırlar)
                cell.fill = input_fill
                cell.border = thick_border
            
            # Hesaplanan değerler
            elif col_idx == 3 and '=' in str(value):
                cell.fill = calc_fill
            
            # Önemli sonuçlar
            elif row_idx in [37, 39, 40] and col_idx == 3:  # Ana sonuçlar (yeni satırlar)
                cell.font = Font(bold=True, size=14)
                cell.fill = result_fill
                cell.border = thick_border
    
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 5
    ws5.column_dimensions['C'].width = 20
    ws5.column_dimensions['D'].width = 15
    
    # VBA Makro kodu oluştur
    vba_code = '''
Sub Worksheet_Change(ByVal Target As Range)
    Application.EnableEvents = False
    Application.ScreenUpdating = False
    
    ' SARI hücrelerde değişiklik olduğunda çalışır
    If Target.Interior.Color = RGB(255, 255, 0) Then
        ' Hesaplama sayfasındaki sonuçları güncelle
        Worksheets("Dashboard").Range("C10").Value = "Güncelleniyor..."
        Application.Calculate
        DoEvents
        
        ' Durum güncelle
        Dim totalCost As Double
        totalCost = Worksheets("Hesaplama").Range("C35").Value
        
        If totalCost > 0 Then
            Worksheets("Dashboard").Range("C10").Value = "Hesaplandı ✓"
        Else
            Worksheets("Dashboard").Range("C10").Value = "Hata ⚠"
        End If
    End If
    
    Application.EnableEvents = True
    Application.ScreenUpdating = True
End Sub

Sub HesaplamaGuncelle()
    Application.Calculate
    MsgBox "Tüm hesaplamalar güncellendi!", vbInformation
End Sub
'''
    
    # VBA modülü eklemek için dosyayı kaydet ve yorum ekle
    filename = 'Sera_Ultimate_Professional_System_v7.xlsx'
    wb.save(filename)
    
    print(f"🚀 ULTIMATE PROFESYONEL SERA SİSTEMİ oluşturuldu: {filename}")
    print("\n✨ PROFESYONEL ÖZELLİKLER:")
    print("🟡 SARI hücreler = INPUT alanları (net belirtildi)")
    print("🔗 Sayfalar arası tam entegrasyon")
    print("📊 Dashboard kontrol paneli")
    print("⚡ Etkileşimli hesaplama sistemi")
    print("💎 5 sekme: Dashboard + 3 Malzeme + Hesaplama")
    print("🎯 Manuel input alanları net olarak belirtildi")
    print("📈 Otomatik güncelleme sistemi")
    print("\n📋 KULLANIM:")
    print("1. Dashboard: Proje genel görünümü")
    print("2-3-4. Malzeme sayfaları: SARI hücrelerde fiyat güncelleyin")
    print("5. Hesaplama: SARI hücrelerde proje boyutlarını girin")
    print("\n🔧 VBA MAKRO KOD (Excel'e elle eklenecek):")
    print("Bu kodu VBA editöründe Worksheet_Change olarak ekleyin:")
    print(vba_code)
    
    return filename

if __name__ == "__main__":
    create_ultimate_professional_excel()
